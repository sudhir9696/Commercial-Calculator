"""CRE Deal Screener — Streamlit dashboard with a Due Diligence Command Center.

Two tabs:
- Screener  — pulls GA listings from Apify (`crawlerbros/crexi-real-estate-scraper`),
              flags 'Action Required' deals, and exposes per-deal tools (OM
              handoff, PropTracer LLC capture, CCIM Excel).
- Analyzer  — paste a Crexi link, upload an OM PDF, or enter fields manually;
              get a CCIM model with a full levered IRR / NPV / equity-multiple
              projection (numpy_financial — same math as tvm.py).
"""
from __future__ import annotations

import io
import json
import os
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import altair as alt
import numpy_financial as npf
import pandas as pd
import requests
import streamlit as st
from dotenv import load_dotenv

# Optional apify-client. The library's v2 release switched return values from
# dicts to Pydantic models, which broke our `["id"]` access. We default to raw
# REST (most reliable across Cloud Python versions and v1/v2 client behavior)
# and let users opt in to apify-client via APIFY_USE_CLIENT=1 once they've
# verified it works in their env.
try:
    from apify_client import ApifyClient
    _APIFY_CLIENT_AVAILABLE = True
    _APIFY_CLIENT_IMPORT_ERROR: str | None = None
except Exception as _exc:
    ApifyClient = None  # type: ignore[assignment]
    _APIFY_CLIENT_AVAILABLE = False
    _APIFY_CLIENT_IMPORT_ERROR = repr(_exc)

_USE_APIFY_CLIENT = _APIFY_CLIENT_AVAILABLE and os.getenv("APIFY_USE_CLIENT", "0") == "1"

# Anthropic SDK is also optional — if the install fails on Cloud or the user
# hasn't set ANTHROPIC_API_KEY, the AI Analyst tab degrades gracefully.
try:
    from anthropic import Anthropic
    _ANTHROPIC_AVAILABLE = True
except Exception as _exc:
    Anthropic = None  # type: ignore[assignment]
    _ANTHROPIC_AVAILABLE = False


def _to_dict(obj: Any) -> dict[str, Any]:
    """Normalize apify-client v2 Pydantic models (or dicts) into a plain dict.

    apify-client v1.x returned plain dicts. v2.x returns Pydantic v2 models,
    which don't support subscript access (`obj["id"]`) and break our existing
    dict-shaped consumer code. This helper unifies both shapes.
    """
    if obj is None:
        return {}
    if isinstance(obj, dict):
        return obj
    for method in ("model_dump", "dict", "to_dict"):
        fn = getattr(obj, method, None)
        if callable(fn):
            try:
                return fn()
            except Exception:
                continue
    if hasattr(obj, "__dict__"):
        return {k: v for k, v in obj.__dict__.items() if not k.startswith("_")}
    return {}
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None  # PDF upload will be disabled with an explanatory message

load_dotenv()


def _secret(name: str, default: str = "") -> str:
    """Read from env first, fall back to st.secrets so the same code works on
    Streamlit Cloud (which exposes values via st.secrets, not env vars)."""
    val = os.getenv(name, "").strip()
    if val:
        return val
    try:
        return str(st.secrets.get(name, default)).strip()
    except (FileNotFoundError, AttributeError, Exception):
        return default


APIFY_TOKEN = _secret("APIFY_TOKEN")
APIFY_DATASET_ID = _secret("APIFY_DATASET_ID")
ANTHROPIC_API_KEY = _secret("ANTHROPIC_API_KEY")
SCRAPINGBEE_API_KEY = _secret("SCRAPINGBEE_API_KEY")
TOKEN_PLACEHOLDER = "your_apify_token_here"
ANTHROPIC_MODEL = "claude-sonnet-4-6"
ANTHROPIC_PLACEHOLDER = "your_anthropic_api_key_here"
SCRAPINGBEE_BASE_URL = "https://app.scrapingbee.com/api/v1/"


def _has_scrapingbee_key() -> bool:
    return bool(SCRAPINGBEE_API_KEY) and not SCRAPINGBEE_API_KEY.lower().startswith("your_")


def fetch_via_scrapingbee(target_url: str, *, render_js: bool = True, country: str = "us") -> str:
    """Hit ScrapingBee's generic scrape endpoint and return the rendered HTML.

    Uses premium residential proxy + JS render so the request looks like a
    real browser to Crexi's Cloudflare layer. Burns 1 credit on the free tier
    (1,000 credits/month included).
    """
    if not _has_scrapingbee_key():
        raise RuntimeError("SCRAPINGBEE_API_KEY not set in env or st.secrets.")
    params = {
        "api_key": SCRAPINGBEE_API_KEY,
        "url": target_url,
        "render_js": "true" if render_js else "false",
        "premium_proxy": "true",
        "country_code": country,
    }
    resp = requests.get(SCRAPINGBEE_BASE_URL, params=params, timeout=120)
    if not resp.ok:
        raise RuntimeError(f"ScrapingBee error {resp.status_code}: {resp.text[:300]}")
    return resp.text


def extract_crexi_listing_urls(html: str) -> list[str]:
    """Pull out unique Crexi property URLs from a search-results page HTML."""
    if not html:
        return []
    # crexi.com/properties/<numeric_id>/<slug>
    pattern = r"https?://www\.crexi\.com/properties/\d+/[a-zA-Z0-9_\-]+"
    urls = sorted({m.split("?")[0] for m in re.findall(pattern, html)})
    return urls

DEFAULT_ACTOR_ID = "skootle~crexi-commercial-real-estate-scraper"
DEFAULT_STATE_CODE = "GA"

# Actor catalog. Skootle is the default because it's the only one we've
# verified returns full-field data (62 fields on a live propertyUrls fetch
# this session). Crawlerbros is included for compatibility but returns only
# ~6 sparse fields per item — the screener's verdict logic will be blind.
ACTOR_CATALOG: dict[str, dict[str, Any]] = {
    "skootle (recommended)": {
        "id": "skootle~crexi-commercial-real-estate-scraper",
        "warning": None,
        "supports_search": True,
        "supports_property_urls": True,
    },
    "crawlerbros (sparse — verdict logic blind)": {
        "id": "crawlerbros~crexi-real-estate-scraper",
        "warning": (
            "⚠️ crawlerbros returns only 6 fields per item (no cap rate, no SF, no address). "
            "Action-Required / GO verdicts can't compute on this data."
        ),
        "supports_search": True,
        "supports_property_urls": False,
    },
}

# All US states (2-letter code), for the State selector. GA is the default.
US_STATES: list[str] = [
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA",
    "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD",
    "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ",
    "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC",
    "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY", "DC",
]

# Sub-class keyword library (case-insensitive scan over title + description +
# native subtype tags). Used to add a `Sub-Class` column in the final table.
# Crexi's native Property Type taxonomy — transcribed from the live UI
# checkboxes as of 2026-06. Top-level keys and sub-category strings match
# Crexi exactly so payload `propertyTypes` lines up with the filter API.
CREXI_TAXONOMY: dict[str, list[str]] = {
    "Retail": [
        "Bank", "Convenience Store", "Day Care/Nursery", "QSR/Fast Food",
        "Gas Station", "Grocery Store", "Pharmacy/Drug", "Restaurant",
        "Bar", "Storefront", "Shopping Center", "Auto Shop",
    ],
    "Multifamily": [
        "Student Housing", "Single Family Rental Portfolio",
        "RV Park", "Apartment Building",
    ],
    "Office": [
        "Traditional Office", "Executive Office", "Medical Office",
        "Creative Office",
    ],
    "Industrial": [
        "Distribution", "Flex", "Warehouse", "R&D",
        "Manufacturing", "Refrigerated/Cold Storage",
    ],
    "Hospitality": [
        "Hotel", "Motel", "Casino",
    ],
    "Mixed Use": [],
    "Land": [
        "Agricultural", "Residential", "Commercial", "Industrial",
        "Islands", "Farm", "Ranch", "Timber", "Hunting/Recreational",
    ],
    "Self Storage": [],
    "Mobile Home Park": [],
    "Senior Living": [],
    "Special Purpose": [
        "Telecom/Data Center", "Sports/Entertainment", "Marina",
        "Golf Course", "School", "Religious/Church", "Garage/Parking",
        "Car Wash", "Airport",
    ],
    "Note/Loan": [],
    "Business for Sale": [],
}

# Crexi sub-types that trigger the 15-yr accelerated depreciation flag
# (IRS Class 57.1) and the Phase 1 environmental advisory.
# Conservative: only definite fuel-canopy + chemical-runoff sub-types tagged
# tax-alpha. Auto Shop is Phase 1 (used motor oil) but NOT 15-yr eligible.
CREXI_SUBTYPE_TAX_ALPHA: set[str] = {"Gas Station", "Car Wash"}
CREXI_SUBTYPE_PHASE_1: set[str] = {"Gas Station", "Car Wash", "Auto Shop"}

SUBCLASS_KEYWORDS: dict[str, list[str]] = {
    "Gas Station": [
        "gas station", " gas ", "fuel", "petroleum", "mpd", "canopy",
        "convenience store", "c-store", " station ", "filling station",
    ],
    "Car Wash": [
        "car wash", " wash ", "tunnel wash", "detail", "wand", "express wash",
        "automated wash",
    ],
    "Laundromat": [
        "laundromat", "coin laundry", "wash-and-fold", "wash and fold",
        "coin operated", "coin-op",
    ],
    "Quick-Lube / Auto Service": [
        "quick lube", "quick-lube", "oil change", "lube center", "auto service",
        "automotive service", "lube + tire",
    ],
    "Self-Storage": [
        "self storage", "self-storage", "mini-storage", "storage facility",
        "climate controlled storage",
    ],
}
DEFAULT_MIN_CAP = 7.0
DEFAULT_MAX_PRICE = 5_000_000
DEFAULT_MAX_PROPERTIES = 10
DEFAULT_RUN_TIMEOUT_SECS = 900
LAST_FETCH_PATH = Path(".streamlit/cache/last_fetch.json")

# UI asset-class label → Crexi propertyTypes array.
# Skootle's actor doesn't currently expose a propertyTypes filter, so the array is
# also future-proofing; the user-friendly label is concatenated into searchKeywords
# (which Crexi's search DOES parse) for real filtering today.
ASSET_CLASS_CATALOG: dict[str, list[str]] = {
    "Industrial Flex": ["Industrial"],
    "Value-Add Multifamily": ["Multifamily"],
    "Retail Motor Fuels Outlet (Gas Station)": ["Retail", "Specialty"],
    "Express Car Wash": ["Retail", "Specialty"],
    "Laundromat / Retail Strip": ["Retail"],
    "Quick-Lube / Automotive Service": ["Retail", "Specialty"],
    "Self-Storage Facility": ["Specialty", "Industrial"],
}

# Ancillary-revenue and distress keywords scanned case-insensitively in each
# listing's description (and title as fallback). Grouped here only for docs;
# every term contributes to the per-deal `flags` list.
ANCILLARY_KEYWORDS: dict[str, list[str]] = {
    "Gas Stations / Laundromats": [
        "COAM", "lottery", "unbranded", "gaming", "vending", "card-operated",
    ],
    "Car Washes / Automotive": [
        "bay", "tunnel", "automated", "equipment package", "replacement cost",
    ],
    "Self-Storage": [
        "climate controlled", "expansion potential", "non-paying tenants", "occupancy upside",
    ],
}
ALL_ANCILLARY_KEYWORDS: list[str] = sorted({k for kws in ANCILLARY_KEYWORDS.values() for k in kws})

# The subset of ancillary terms that, when present, elevate the verdict to
# 🟢 GO (High-Tax-Alpha Asset) — i.e. high-margin secondary revenue or upside
# strong enough to override the cap/price screen.
HIGH_MARGIN_TRIGGERS = {
    "COAM", "lottery", "unbranded", "gaming",
    "expansion potential", "occupancy upside",
}

# IRS Class 57.1 — qualifies for 15-year accelerated depreciation
# (vs. the 39-year default for non-residential real property).
TAX_ALPHA_ASSET_CLASSES = {
    "Retail Motor Fuels Outlet (Gas Station)",
    "Express Car Wash",
}

# Asset classes that require a Phase 1 environmental audit (USTs for fuel,
# chemical runoff for car wash, used motor oil for quick-lube).
PHASE_1_ENV_ASSET_CLASSES = {
    "Retail Motor Fuels Outlet (Gas Station)",
    "Express Car Wash",
    "Quick-Lube / Automotive Service",
}

# CCIM-style underwriting defaults — matches the conventions used in tvm.py.
DEFAULT_HOLD_YEARS = 5
DEFAULT_NOI_GROWTH = 3.0
DEFAULT_EXIT_CAP_DELTA_BPS = 50
DEFAULT_LTV = 65.0
DEFAULT_LOAN_RATE = 7.0
DEFAULT_AMORT_YEARS = 25
DEFAULT_DISCOUNT_RATE = 10.0

# Per Product Owner spec: no mock/sample data. Empty state shown until a
# live Apify run completes.

GA_ALIASES = {"GA", "GEORGIA", "GA.", "GA,"}


# ---------- Coercion ----------

def _coerce_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    if not cleaned or cleaned in {"-", ".", "-."}:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _coerce_int(value: Any) -> int | None:
    f = _coerce_float(value)
    return int(f) if f is not None and not pd.isna(f) else None


def _first_present(row: dict, keys: Iterable[str]) -> Any:
    for k in keys:
        if k in row and row[k] not in (None, ""):
            return row[k]
    return None


# ---------- Address / GA / OM extraction ----------

def _compose_address(row: dict) -> str:
    line1 = _first_present(row, ("address", "street", "propertyAddress")) or ""
    city = row.get("city") or ""
    state = row.get("state") or ""
    zip_code = row.get("zip") or row.get("zip_code") or ""
    parts = [str(line1).strip(), ", ".join(p for p in (city, state) if p).strip(), str(zip_code).strip()]
    composed = ", ".join(p for p in parts if p)
    if composed:
        return composed
    # Skootle frequently puts the address in `title` ("601 Atlanta Rd, Cumming, GA 30040")
    # when the dedicated `address` field is null. Prefer the title in that case.
    return str(row.get("title") or row.get("name") or "—")


def _extract_om_url(row: dict) -> str:
    docs = row.get("documents")
    if isinstance(docs, list):
        candidates: list[tuple[int, str]] = []
        for d in docs:
            if isinstance(d, str):
                candidates.append((1 if d.lower().endswith(".pdf") else 2, d))
            elif isinstance(d, dict):
                url = d.get("url") or d.get("href") or d.get("link") or ""
                name = (d.get("name") or d.get("title") or "").lower()
                if not url:
                    continue
                if any(k in name for k in ("offering", "memorandum", " om", "brochure", "flyer")):
                    candidates.append((0, url))
                elif url.lower().endswith(".pdf"):
                    candidates.append((1, url))
                else:
                    candidates.append((2, url))
        if candidates:
            candidates.sort(key=lambda c: c[0])
            return candidates[0][1]
    for key in ("om_url", "flyerUrl", "brochureUrl", "omUrl", "offeringMemorandumUrl"):
        if row.get(key):
            return str(row[key])
    return ""


def _is_georgia(row: dict) -> bool:
    """Decide if a listing is actually in Georgia.

    Skootle occasionally tags rows with `state: GA` even when the address is
    out-of-state (likely because it parsed the broker's address instead of the
    property's, then defaulted state from the search). So we trust the address
    body over the state field: if the address contains a `<state> <zip>` pattern
    where state isn't GA, reject.
    """
    state = str(row.get("state") or "").strip().upper().rstrip(".,")
    addr_blob = " ".join(
        str(row.get(k) or "") for k in
        ("address", "title", "propertyAddress", "fullAddress", "city")
    ).upper()

    # Find any "XX 12345" patterns (state code before a 5-digit zip).
    state_zip_matches = re.findall(r"\b([A-Z]{2})\s+\d{5}(?:-\d{4})?\b", addr_blob)
    if state_zip_matches:
        # If every state-zip pair is GA, accept. If any is non-GA, reject.
        if all(s in GA_ALIASES for s in state_zip_matches):
            return True
        return False

    if state in GA_ALIASES:
        return True
    return bool(re.search(r"\b(GA|GEORGIA)\b", addr_blob))


def extract_subclass(raw: dict, fallback: str = "") -> str:
    """Identify the asset sub-class by scanning title + description + native subtype tags.

    Priority order:
      1. Match keyword groups in SUBCLASS_KEYWORDS (Gas Station, Car Wash, etc.).
      2. Use the actor's native subtype field (`propertySubType`, `assetSubType`)
         when no keyword hits.
      3. Fall back to the main `assetClass` / `property_type` field.
    """
    blob = " ".join(
        str(raw.get(k) or "")
        for k in (
            "title", "description", "agentMarkdown", "name",
            "propertySubType", "assetSubType", "tenantName",
            "address",
        )
    ).lower()
    if blob:
        for subclass, keywords in SUBCLASS_KEYWORDS.items():
            if any(kw.lower() in blob for kw in keywords):
                return subclass
    native_subtype = raw.get("propertySubType") or raw.get("assetSubType")
    if native_subtype:
        return str(native_subtype)
    return fallback or ""


def normalize_rows(rows: list[dict], *, georgia_only: bool = True) -> pd.DataFrame:
    """Normalize Apify rows into the dashboard's canonical schema.

    Skootle prefers numeric mirrors (`askingPriceUsd`, `capRatePct`, `buildingSqft`);
    the `_first_present` chains keep older field names from crawlerbros et al.
    working as fallbacks.

    `georgia_only` drops rows that have explicit non-GA state info; rows missing
    state info are kept (the actor was already filtered to GA, so trust that).
    """
    canonical: list[dict[str, Any]] = []
    for r in rows:
        explicit_state = str(r.get("state") or "").strip().upper().rstrip(".,")
        # Always defer to _is_georgia — it handles the Skootle mis-tag case
        # (state="GA" on an out-of-state property) by reading the address body.
        if georgia_only and not _is_georgia(r):
            continue
        property_type = _first_present(r, ("assetClass", "assetSubType", "propertySubType", "propertyType", "property_type")) or ""
        canonical.append({
            "address": _compose_address(r),
            "state": explicit_state,
            "property_type": property_type,
            "sub_class": extract_subclass(r, fallback=property_type),
            "transaction_type_native": str(_first_present(r, ("transactionType", "transaction_type", "listingType", "type")) or "").lower(),
            "asking_price": _coerce_float(_first_present(r, ("askingPriceUsd", "askingPrice", "price", "asking_price", "listPrice"))),
            "cap_rate_pct": _coerce_float(_first_present(r, ("capRatePct", "capRate", "cap_rate", "cap_rate_pct"))),
            "square_footage": _coerce_int(_first_present(r, ("squareFootageNum", "squareFootage", "buildingSqft", "square_footage", "squareFeet", "buildingSize", "sf", "size"))),
            "om_url": _extract_om_url(r),
            "listing_url": _first_present(r, ("listingUrl", "url", "property_url", "listing_url", "detailPageUrl")) or "",
            "description": _first_present(r, ("description", "agentMarkdown", "highlights")) or "",
        })
    return pd.DataFrame(canonical)


# ---------- Apify integration ----------

def _apify_error_message(resp: requests.Response) -> str:
    try:
        body = resp.json()
        msg = body.get("error", {}).get("message") or body.get("message") or resp.text[:400]
    except ValueError:
        msg = resp.text[:400]
    return f"HTTP {resp.status_code}: {msg}"


# ---------- Last-fetch persistence (survives session resets within a container) ----------

def save_last_fetch(rows: list[dict], source: str, query: str) -> None:
    try:
        LAST_FETCH_PATH.parent.mkdir(parents=True, exist_ok=True)
        LAST_FETCH_PATH.write_text(json.dumps({
            "rows": rows,
            "source": source,
            "query": query,
            "ts": datetime.utcnow().isoformat() + "Z",
        }))
    except Exception:
        pass


def load_last_fetch() -> dict | None:
    try:
        if not LAST_FETCH_PATH.exists():
            return None
        return json.loads(LAST_FETCH_PATH.read_text())
    except Exception:
        return None


def clear_last_fetch() -> None:
    try:
        if LAST_FETCH_PATH.exists():
            LAST_FETCH_PATH.unlink()
    except Exception:
        pass


def _build_actor_payload(
    *,
    search_keywords: list[str] | None,
    start_urls: list[str] | None,
    property_urls: list[str] | None,
    property_types: list[str] | None,
    locations: list[str] | None,
    max_items: int,
    max_search_pages: int,
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "maxItems": int(max_items),
        "maxSearchPages": int(max_search_pages),
        "proxyConfiguration": {"useApifyProxy": True, "apifyProxyGroups": ["RESIDENTIAL"]},
    }
    if search_keywords:
        payload["searchKeywords"] = search_keywords
    if start_urls:
        payload["startUrls"] = [{"url": u} for u in start_urls]
    if property_urls:
        payload["propertyUrls"] = property_urls
    if property_types:
        # Forward-compat: skootle currently ignores propertyTypes; included so
        # the asset-class catalog mapping is exercised end-to-end.
        payload["propertyTypes"] = property_types
    if locations:
        # Forward-compat: skootle has no `locations` input today. Crawlerbros
        # accepts state codes here only. We pass the user's City/County so the
        # mapping is documented in code; the actual filtering happens via
        # `searchKeywords` which Crexi's free-text search does parse.
        payload["locations"] = locations
    return payload


# ---------- Apify backend abstraction (apify-client preferred, REST fallback) ----------

def _apify_start_run(
    token: str, actor_id: str, payload: dict[str, Any], *,
    actor_timeout_secs: int = DEFAULT_RUN_TIMEOUT_SECS,
) -> str:
    """Start a run and set the platform-level max run time.

    Apify's default actor run time is conservatively short (~5 min for many
    actors). Crexi scraping over a residential proxy regularly needs more —
    we pass `actor_timeout_secs` (default 900 = 15 min) so the platform
    doesn't kill the run before it finishes scraping the first page.
    """
    if _USE_APIFY_CLIENT and ApifyClient is not None:
        try:
            run = _to_dict(ApifyClient(token).actor(actor_id).start(
                run_input=payload, timeout_secs=actor_timeout_secs,
            ))
            run_id = run.get("id")
            if run_id:
                return str(run_id)
        except Exception:
            pass  # fall through to REST
    url = f"https://api.apify.com/v2/acts/{actor_id}/runs"
    resp = requests.post(
        url,
        params={"token": token, "timeout": actor_timeout_secs},
        json=payload,
        timeout=30,
    )
    if not resp.ok:
        raise RuntimeError(_apify_error_message(resp))
    return resp.json()["data"]["id"]


def _apify_get_run(token: str, run_id: str) -> dict[str, Any]:
    if _USE_APIFY_CLIENT and ApifyClient is not None:
        run = _to_dict(ApifyClient(token).run(run_id).get())
        if run:
            return run
    url = f"https://api.apify.com/v2/actor-runs/{run_id}"
    resp = requests.get(url, params={"token": token}, timeout=30)
    if not resp.ok:
        raise RuntimeError(_apify_error_message(resp))
    return resp.json()["data"]


def _apify_abort_run(token: str, run_id: str) -> None:
    try:
        if _USE_APIFY_CLIENT and ApifyClient is not None:
            ApifyClient(token).run(run_id).abort()
            return
    except Exception:
        pass
    try:
        requests.post(
            f"https://api.apify.com/v2/actor-runs/{run_id}/abort",
            params={"token": token}, timeout=15,
        )
    except Exception:
        pass


def _apify_list_dataset(token: str, dataset_id: str) -> list[dict]:
    if _USE_APIFY_CLIENT and ApifyClient is not None:
        try:
            items = list(ApifyClient(token).dataset(dataset_id).iterate_items())
            return [_to_dict(it) if not isinstance(it, dict) else it for it in items]
        except Exception:
            pass  # fall through to REST
    url = f"https://api.apify.com/v2/datasets/{dataset_id}/items"
    resp = requests.get(url, params={"token": token, "format": "json"}, timeout=60)
    if not resp.ok:
        raise RuntimeError(_apify_error_message(resp))
    return resp.json()


def run_actor_async(
    token: str,
    actor_id: str,
    *,
    search_keywords: list[str] | None = None,
    start_urls: list[str] | None = None,
    property_urls: list[str] | None = None,
    property_types: list[str] | None = None,
    locations: list[str] | None = None,
    max_items: int = 10,
    max_search_pages: int = 5,
    timeout_secs: int = DEFAULT_RUN_TIMEOUT_SECS,
    poll_seconds: int = 5,
    progress_cb=None,
) -> tuple[list[dict], dict[str, Any]]:
    """Start an Apify actor run, poll for progress, return its dataset items.

    Routes the underlying calls through apify-client when available, falls back
    to raw REST. The explicit poll loop is kept so the Streamlit UI can show
    live progress via `progress_cb` (apify-client's blocking `.call()` would
    freeze the UI with no feedback).
    """
    payload = _build_actor_payload(
        search_keywords=search_keywords, start_urls=start_urls, property_urls=property_urls,
        property_types=property_types, locations=locations,
        max_items=max_items, max_search_pages=max_search_pages,
    )
    run_id = _apify_start_run(token, actor_id, payload, actor_timeout_secs=timeout_secs)
    start = time.monotonic()
    last: dict[str, Any] = {"id": run_id, "status": "READY"}

    while True:
        last = _apify_get_run(token, run_id)
        elapsed = int(time.monotonic() - start)
        item_count = (last.get("stats") or {}).get("itemCount", 0)
        if progress_cb:
            progress_cb(elapsed, last.get("status", "?"), item_count)
        if last.get("status") in ("SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"):
            break
        if elapsed >= timeout_secs:
            _apify_abort_run(token, run_id)
            last["status"] = "ABORTED"
            last["statusMessage"] = f"Client-side timeout after {timeout_secs}s."
            break
        time.sleep(poll_seconds)

    final_status = last.get("status")
    dataset_id = last.get("defaultDatasetId")

    if final_status == "SUCCEEDED":
        if not dataset_id:
            return [], last
        return _apify_list_dataset(token, dataset_id), last

    # Non-success status — try to recover partial results before raising.
    # On TIMED-OUT / ABORTED, the actor often saved a portion of the dataset
    # before being killed. Returning those lets the user see SOMETHING instead
    # of a hard error after burning proxy credits.
    if dataset_id:
        try:
            partial_items = _apify_list_dataset(token, dataset_id)
        except Exception:
            partial_items = []
        if partial_items:
            last["_partial"] = True
            last["_partial_count"] = len(partial_items)
            return partial_items, last

    msg = last.get("statusMessage") or final_status or "UNKNOWN"
    raise RuntimeError(f"Apify run ended with status {final_status}: {msg}")


def run_actor_sync(
    token: str, actor_id: str, *,
    search_keywords: list[str] | None = None,
    start_urls: list[str] | None = None,
    property_urls: list[str] | None = None,
    property_types: list[str] | None = None,
    locations: list[str] | None = None,
    max_items: int = 10,
    max_search_pages: int = 5,
    timeout_secs: int = 600,
) -> list[dict]:
    """Synchronous single-shot — used by the analyzer's single-URL fetch.

    Uses apify-client's blocking `.call()` when available; otherwise hits the
    REST `/run-sync-get-dataset-items` endpoint directly.
    """
    payload = _build_actor_payload(
        search_keywords=search_keywords, start_urls=start_urls, property_urls=property_urls,
        property_types=property_types, locations=locations,
        max_items=max_items, max_search_pages=max_search_pages,
    )
    if _USE_APIFY_CLIENT and ApifyClient is not None:
        try:
            run = _to_dict(
                ApifyClient(token).actor(actor_id).call(run_input=payload, timeout_secs=timeout_secs)
            )
            if run and run.get("status") == "SUCCEEDED":
                dataset_id = run.get("defaultDatasetId")
                return _apify_list_dataset(token, dataset_id) if dataset_id else []
            # If we got a non-success status, fall through to REST so the user
            # at least gets a clean error from the canonical endpoint.
        except Exception:
            pass  # fall through to REST
    url = f"https://api.apify.com/v2/acts/{actor_id}/run-sync-get-dataset-items"
    resp = requests.post(
        url, params={"token": token, "timeout": timeout_secs, "format": "json"},
        json=payload, timeout=timeout_secs + 30,
    )
    if not resp.ok:
        raise RuntimeError(_apify_error_message(resp))
    return resp.json()


@st.cache_data(ttl=600, show_spinner=False)
def fetch_dataset_items(token: str, dataset_id: str) -> list[dict]:
    """Load an existing Apify dataset by id (used when the user pastes one)."""
    return _apify_list_dataset(token, dataset_id)


# ---------- Flagging ----------

VERDICT_GO = "🟢 GO (High-Tax-Alpha Asset)"
VERDICT_ACTION = "🟢 Action Required"
VERDICT_REVIEW = "⚪ Review"


def _scan_description_for_flags(text: Any) -> list[str]:
    if not isinstance(text, str) or not text:
        return []
    low = text.lower()
    return [kw for kw in ALL_ANCILLARY_KEYWORDS if kw.lower() in low]


def analyze_and_score(df: pd.DataFrame, sb: dict[str, Any]) -> pd.DataFrame:
    """CCIM-style scoring: keyword flags + 15-yr depreciation tag + verdict cascade.

    Verdict cascade (last winner takes the row):
      1. ⚪ Review  ← default
      2. 🟢 Action Required  ← cap >= min_cap AND price <= max_price_rule
      3. 🟢 GO (High-Tax-Alpha Asset)  ← either condition:
         - sidebar asset class qualifies for IRS Class 57.1 (15-yr accel.
           depreciation) AND the listing has a real price, OR
         - the listing's description contains any HIGH_MARGIN_TRIGGER
           keyword (COAM, lottery, unbranded, gaming, expansion potential,
           occupancy upside) — signals disproportionate ancillary upside.
    """
    df = df.copy()
    if df.empty:
        df["flags"] = pd.Series([], dtype=object)
        df["15_Yr_Accelerated_Depreciation"] = pd.Series([], dtype=bool)
        df["verdict"] = pd.Series([], dtype=str)
        return df

    desc_col = df["description"] if "description" in df.columns else df["address"].fillna("")
    title_col = df.get("title", pd.Series([""] * len(df)))
    combined_text = desc_col.fillna("") + " " + title_col.fillna("")
    df["flags"] = combined_text.apply(_scan_description_for_flags)

    asset_label = (sb.get("asset_class") or "").strip()
    crexi_sub = (sb.get("crexi_sub_type") or "").strip()
    is_tax_alpha_class = (
        asset_label in TAX_ALPHA_ASSET_CLASSES
        or crexi_sub in CREXI_SUBTYPE_TAX_ALPHA
    )
    df["15_Yr_Accelerated_Depreciation"] = bool(is_tax_alpha_class)

    df["verdict"] = VERDICT_REVIEW

    cap_ok = df["cap_rate_pct"].fillna(-1).ge(sb["min_cap"])
    price_ok = df["asking_price"].fillna(float("inf")).le(sb["max_price_rule"])
    df.loc[cap_ok & price_ok, "verdict"] = VERDICT_ACTION

    has_price = df["asking_price"].notna() & df["asking_price"].gt(0)
    has_high_margin = df["flags"].apply(lambda fs: bool(set(fs) & HIGH_MARGIN_TRIGGERS))
    if is_tax_alpha_class:
        df.loc[has_price, "verdict"] = VERDICT_GO
    df.loc[has_high_margin, "verdict"] = VERDICT_GO

    return df


# ---------- Analyzer helpers: URL + PDF + investment math ----------

def extract_crexi_property_id(url: str) -> str | None:
    m = re.search(r"crexi\.com/properties/(\d+)", url or "")
    return m.group(1) if m else None


def extract_crexi_slug_hint(url: str) -> str:
    """The Crexi URL slug often encodes 'name-city-state' — useful as a fallback hint."""
    m = re.search(r"crexi\.com/properties/\d+/([a-z0-9\-]+)", url or "", re.I)
    if not m:
        return ""
    return m.group(1).replace("-", " ").title()


def extract_pdf_text(file_bytes: bytes) -> str:
    if PdfReader is None:
        return ""
    try:
        reader = PdfReader(io.BytesIO(file_bytes))
        return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception:
        return ""


def parse_om_fields(text: str) -> dict[str, Any]:
    """Best-effort regex extraction of price / cap rate / SF / NOI from OM text."""
    out: dict[str, Any] = {}
    if not text:
        return out
    flat = re.sub(r"\s+", " ", text)

    m = re.search(r"(?:offering|asking|list(?:ing)?|sale)\s*price[^$]{0,30}\$\s?([\d,]+(?:\.\d+)?)", flat, re.I)
    if not m:
        m = re.search(r"price[^$]{0,15}\$\s?([\d,]+(?:\.\d+)?)", flat, re.I)
    if m:
        out["price"] = float(m.group(1).replace(",", ""))

    m = re.search(r"(?:going[-\s]?in\s*)?cap(?:italization)?\s*rate[^%]{0,40}(\d{1,2}(?:\.\d+)?)\s?%", flat, re.I)
    if m:
        out["cap_rate"] = float(m.group(1))

    m = re.search(
        r"(?:building|total|gross|rentable|net\s*rentable|net)\s*(?:square\s*feet|sq\.?\s?ft\.?|sf|area|size)[^\d]{0,30}([\d,]+)",
        flat,
        re.I,
    )
    if not m:
        m = re.search(r"([\d,]{4,})\s*(?:sf|sq\.?\s?ft\.?|square\s?feet)\b", flat, re.I)
    if m:
        out["sf"] = int(m.group(1).replace(",", ""))

    m = re.search(r"(?:noi|net\s*operating\s*income)[^$]{0,40}\$\s?([\d,]+(?:\.\d+)?)", flat, re.I)
    if m:
        out["noi"] = float(m.group(1).replace(",", ""))
    return out


def project_investment(
    *,
    price: float,
    cap_rate_pct: float,
    hold_years: int,
    noi_growth_pct: float,
    exit_cap_pct: float,
    ltv_pct: float,
    loan_rate_pct: float,
    amort_years: int,
    discount_rate_pct: float,
    noi_override: float | None = None,
) -> dict[str, Any]:
    """Full levered DCF projection mirroring tvm.py's numpy_financial conventions.

    Returns the assumptions, year-by-year cash flow series, and key return metrics
    (CoC y1, unlevered/levered IRR, NPV @ discount rate, equity multiple).
    """
    price = float(price or 0.0)
    cap = (cap_rate_pct or 0.0) / 100.0
    growth = (noi_growth_pct or 0.0) / 100.0
    exit_cap = (exit_cap_pct or 0.0) / 100.0
    ltv = (ltv_pct or 0.0) / 100.0
    loan_rate = (loan_rate_pct or 0.0) / 100.0
    discount = (discount_rate_pct or 0.0) / 100.0
    years = max(int(hold_years or 1), 1)
    amort = max(int(amort_years or 1), 1)

    loan = price * ltv
    equity = price - loan
    noi_y1 = float(noi_override) if noi_override else price * cap

    monthly_rate = loan_rate / 12.0 if loan_rate else 0.0
    months = amort * 12
    if loan > 0 and monthly_rate > 0:
        monthly_pmt = float(-npf.pmt(monthly_rate, months, loan))
    elif loan > 0:
        monthly_pmt = loan / months
    else:
        monthly_pmt = 0.0
    annual_ds = monthly_pmt * 12

    nois = [noi_y1 * ((1 + growth) ** (y - 1)) for y in range(1, years + 1)]
    cf_after_debt = [n - annual_ds for n in nois]

    noi_year_after = noi_y1 * ((1 + growth) ** years)
    reversion = (noi_year_after / exit_cap) if exit_cap > 0 else 0.0

    hold_months = years * 12
    if loan > 0 and monthly_rate > 0:
        factor = (1 + monthly_rate) ** hold_months
        loan_balance = loan * factor - monthly_pmt * (factor - 1) / monthly_rate
    elif loan > 0:
        loan_balance = max(loan - monthly_pmt * hold_months, 0.0)
    else:
        loan_balance = 0.0
    loan_balance = max(loan_balance, 0.0)
    net_sale_proceeds = max(reversion - loan_balance, 0.0)

    unlevered_cfs = [-price] + nois[:-1] + [nois[-1] + reversion]
    levered_cfs = [-equity] + cf_after_debt[:-1] + [cf_after_debt[-1] + net_sale_proceeds]

    def _safe_irr(series):
        try:
            v = npf.irr(series)
            return float(v) if v == v else None  # filter NaN
        except Exception:
            return None

    coc_y1 = (nois[0] - annual_ds) / equity if equity > 0 else 0.0
    total_levered_inflows = sum(cf_after_debt) + net_sale_proceeds
    equity_multiple = total_levered_inflows / equity if equity > 0 else 0.0

    return {
        "price": price,
        "loan": loan,
        "equity": equity,
        "noi_y1": noi_y1,
        "monthly_pmt": monthly_pmt,
        "annual_ds": annual_ds,
        "reversion": reversion,
        "loan_balance_at_sale": loan_balance,
        "net_sale_proceeds": net_sale_proceeds,
        "nois": nois,
        "cf_after_debt": cf_after_debt,
        "unlevered_cfs": unlevered_cfs,
        "levered_cfs": levered_cfs,
        "unlevered_irr": _safe_irr(unlevered_cfs),
        "levered_irr": _safe_irr(levered_cfs),
        "levered_npv": float(npf.npv(discount, levered_cfs)),
        "coc_y1": coc_y1,
        "equity_multiple": equity_multiple,
        "assumptions": {
            "hold_years": years, "noi_growth_pct": noi_growth_pct, "exit_cap_pct": exit_cap_pct,
            "ltv_pct": ltv_pct, "loan_rate_pct": loan_rate_pct, "amort_years": amort_years,
            "discount_rate_pct": discount_rate_pct,
        },
    }


def investment_cf_dataframe(inv: dict) -> pd.DataFrame:
    years = list(range(1, len(inv["nois"]) + 1))
    df = pd.DataFrame({
        "Year": years,
        "NOI": inv["nois"],
        "Debt Service": [inv["annual_ds"]] * len(years),
        "CF After Debt": inv["cf_after_debt"],
    })
    df.loc[len(df) - 1, "CF After Debt"] = inv["cf_after_debt"][-1] + inv["net_sale_proceeds"]
    df.loc[len(df) - 1, "Note"] = "+ net sale proceeds"
    return df


# ---------- CCIM Excel ----------

def _slugify(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")[:60] or "deal"


def build_ccim_workbook(deal: dict[str, Any], investment: dict | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Underwriting Summary"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    section_fill = PatternFill("solid", fgColor="1F4E78")
    label_fill = PatternFill("solid", fgColor="F2F2F2")
    section_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B3D62")
    bold = Font(bold=True)
    money = '"$"#,##0.00'
    pct = "0.00%"

    ws.merge_cells("A1:B1")
    ws["A1"] = "CRE Underwriting Summary (CCIM-style)"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A3"], ws["B3"] = "Property", deal.get("address", "—")
    ws["A4"], ws["B4"] = "Listing URL", deal.get("listing_url", "")
    ws["A5"], ws["B5"] = "Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    for r in (3, 4, 5):
        ws.cell(row=r, column=1).font = bold

    def section(row: int, title: str) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row=row, column=1, value=title)
        c.font = section_font
        c.fill = section_fill
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20

    def kv(row, label, value=None, fmt=None, formula=None):
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=1).fill = label_fill
        cell = ws.cell(row=row, column=2, value=(formula if formula is not None else value))
        if fmt:
            cell.number_format = fmt
        cell.border = border

    price = _coerce_float(deal.get("asking_price"))
    cap = _coerce_float(deal.get("cap_rate_pct"))
    sf = _coerce_int(deal.get("square_footage"))

    section(7, "ACQUISITION")
    kv(8, "Asking Price", price, money)
    kv(9, "Cap Rate", (cap / 100.0) if cap is not None else None, pct)
    kv(10, "NOI (Price × Cap)", fmt=money, formula="=B8*B9" if price and cap else None)
    kv(11, "Square Footage", sf, "#,##0")
    kv(12, "Price / SF", fmt=money, formula="=B8/B11" if price and sf else None)

    section(14, "INCOME — fill in for ChatGPT")
    kv(15, "Gross Potential Rent", fmt=money)
    kv(16, "Vacancy", fmt=pct)
    kv(17, "Effective Gross Income", fmt=money, formula="=B15*(1-B16)")

    section(19, "EXPENSES — fill in for ChatGPT")
    kv(20, "Property Taxes", fmt=money)
    kv(21, "Insurance", fmt=money)
    kv(22, "Repairs & Maintenance", fmt=money)
    kv(23, "Management Fee", fmt=money)
    kv(24, "Total OpEx", fmt=money, formula="=SUM(B20:B23)")

    section(26, "RETURNS")
    kv(27, "NOI (EGI − OpEx)", fmt=money, formula="=B17-B24")
    kv(28, "Implied Cap (NOI / Price)", fmt=pct, formula="=IFERROR(B27/B8,0)")
    kv(29, "Cash-on-Cash (post-debt)", fmt=pct)

    section(31, "NOTES — paste analyst / ChatGPT commentary below")
    ws.merge_cells("A32:B40")
    ws["A32"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A32"].border = border

    ws.column_dimensions[get_column_letter(1)].width = 34
    ws.column_dimensions[get_column_letter(2)].width = 42

    if investment:
        _add_investment_sheet(wb, investment)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_investment_sheet(wb: Workbook, inv: dict) -> None:
    ws = wb.create_sheet("Investment Analysis")
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="0B3D62")
    header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    money = '"$"#,##0'
    pct = "0.00%"

    ws.merge_cells("A1:G1")
    ws["A1"] = "Investment Analysis — Levered Projection"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    assumptions = inv["assumptions"]
    rows = [
        ("Price", inv["price"], money),
        ("Loan", inv["loan"], money),
        ("Equity", inv["equity"], money),
        ("Year-1 NOI", inv["noi_y1"], money),
        ("Annual Debt Service", inv["annual_ds"], money),
        ("Hold Years", assumptions["hold_years"], "0"),
        ("NOI Growth", assumptions["noi_growth_pct"] / 100.0, pct),
        ("Exit Cap", assumptions["exit_cap_pct"] / 100.0, pct),
        ("LTV", assumptions["ltv_pct"] / 100.0, pct),
        ("Loan Rate", assumptions["loan_rate_pct"] / 100.0, pct),
        ("Amort Years", assumptions["amort_years"], "0"),
        ("Discount Rate (target yield)", assumptions["discount_rate_pct"] / 100.0, pct),
    ]
    for i, (label, value, fmt) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = bold
        c = ws.cell(row=i, column=2, value=value)
        c.number_format = fmt

    cf_start = len(rows) + 5
    headers = ["Year", "NOI", "Debt Service", "CF After Debt", "Reversion", "Total CF"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=cf_start, column=j, value=h)
        c.font = bold
        c.fill = PatternFill("solid", fgColor="DDEBF7")

    years = list(range(1, len(inv["nois"]) + 1))
    for k, year in enumerate(years):
        r = cf_start + 1 + k
        ws.cell(row=r, column=1, value=year)
        ws.cell(row=r, column=2, value=inv["nois"][k]).number_format = money
        ws.cell(row=r, column=3, value=-inv["annual_ds"]).number_format = money
        ws.cell(row=r, column=4, value=inv["cf_after_debt"][k]).number_format = money
        reversion = inv["net_sale_proceeds"] if year == years[-1] else 0
        ws.cell(row=r, column=5, value=reversion).number_format = money
        total = inv["cf_after_debt"][k] + reversion
        ws.cell(row=r, column=6, value=total).number_format = money

    eq_row = cf_start + len(years) + 2
    ws.cell(row=eq_row, column=1, value="Initial Equity (Year 0)").font = bold
    ws.cell(row=eq_row, column=2, value=-inv["equity"]).number_format = money

    metrics_row = eq_row + 2
    def _metric(r, label, value, fmt):
        ws.cell(row=r, column=1, value=label).font = bold
        c = ws.cell(row=r, column=2, value=value)
        c.number_format = fmt

    _metric(metrics_row, "Cash-on-Cash (Y1)", inv["coc_y1"], pct)
    _metric(metrics_row + 1, "Unlevered IRR", inv["unlevered_irr"] or 0, pct)
    _metric(metrics_row + 2, "Levered IRR", inv["levered_irr"] or 0, pct)
    _metric(metrics_row + 3, "Levered NPV @ discount rate", inv["levered_npv"], money)
    _metric(metrics_row + 4, "Equity Multiple", inv["equity_multiple"], "0.00\"x\"")

    for col, w in enumerate([28, 18, 18, 18, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------- Sidebar (shared) ----------

def _token_ok() -> bool:
    return bool(APIFY_TOKEN) and APIFY_TOKEN != TOKEN_PLACEHOLDER


def render_sidebar() -> dict[str, Any]:
    with st.sidebar:
        st.subheader("🔍 Deal Screener filters")
        st.caption("Filters for the Apify Crexi pull. Underwriting math + screening rule "
                   "moved to the Deal Screener tab.")

        actor_choice = st.selectbox(
            "Apify actor",
            list(ACTOR_CATALOG.keys()),
            index=0, key="sb_actor",
            help="skootle is the only actor that returns full-field data.",
        )
        _actor_meta = ACTOR_CATALOG[actor_choice]
        if _actor_meta["warning"]:
            st.warning(_actor_meta["warning"], icon="⚠️")

        transaction_type = st.radio(
            "Looking for",
            ["For Sale", "For Lease", "Both"],
            index=0, horizontal=True,
            key="sb_txn_type",
            help="Crexi treats sale and lease as separate searches. Choose one for "
                 "tighter results; pick 'Both' if you don't care.",
        )

        state_code = st.selectbox(
            "State", US_STATES,
            index=US_STATES.index(DEFAULT_STATE_CODE),
            key="sb_state",
        )

        crexi_cat = st.selectbox(
            "Crexi category", ["(any)"] + list(CREXI_TAXONOMY.keys()),
            index=0, key="sb_crexi_cat",
            help="Matches Crexi's Property Type filter.",
        )
        crexi_subs_available = CREXI_TAXONOMY.get(crexi_cat, []) if crexi_cat != "(any)" else []
        if crexi_subs_available:
            crexi_sub = st.selectbox(
                "Sub-type", ["(any)"] + crexi_subs_available,
                index=0, key="sb_crexi_sub",
            )
        else:
            crexi_sub = "(any)"

        city_or_county_in = st.text_input(
            "City or County (optional)",
            placeholder="e.g. Atlanta, Hall County",
            key="sb_city",
        )

        pc1, pc2 = st.columns(2)
        with pc1:
            price_min_in = st.number_input(
                "Min price ($M)", 0.0, 500.0, 0.0, 0.25, format="%.2f",
                key="sb_price_min",
            )
        with pc2:
            price_max_in = st.number_input(
                "Max price ($M)", 0.0, 500.0, 0.0, 0.25, format="%.2f",
                key="sb_price_max",
            )
        cap_min_in = st.number_input(
            "Min cap rate (%)", 0.0, 20.0, 0.0, 0.25, format="%.2f",
            key="sb_cap_min",
        )
        lease_type_in = st.selectbox(
            "Lease type",
            ["(any)", "NNN", "NN", "Gross", "Modified Gross", "Absolute Net"],
            index=0, key="sb_lease_type",
        )
        search_keywords_in = st.text_input(
            "Extra keywords (optional)",
            placeholder="e.g. value-add, drive-thru",
            key="sb_keywords",
        )

        # Strategy preset removed — Crexi category + sub-type cover the same
        # ground more precisely. asset_class is kept in the returned dict (None)
        # so downstream tax-alpha checks degrade gracefully.
        asset_class = "(any)"

        # Prepend the transaction type so Crexi's search parser narrows results.
        _txn_query = {
            "For Sale": "for sale",
            "For Lease": "for lease",
            "Both": None,
        }.get(transaction_type)
        _query_parts = [
            _txn_query,
            None if crexi_sub == "(any)" else crexi_sub,
            None if crexi_cat == "(any)" or crexi_sub != "(any)" else crexi_cat,
            city_or_county_in.strip() or None,
            None if lease_type_in == "(any)" else lease_type_in,
            f"${price_min_in:.1f}M-${price_max_in:.1f}M" if (price_min_in or price_max_in) else None,
            f"{cap_min_in:.1f}% cap" if cap_min_in else None,
            search_keywords_in.strip() or None,
            state_code,
        ]
        _preview_query = " ".join(p for p in _query_parts if p)
        st.caption(f"🔎 Query preview: `{_preview_query}`")
        # Auto-discover URLs via ScrapingBee. Free tier (1k credits/month) is
        # enough for ~20 search-page discoveries. Works alongside the manual
        # Bulk URL paste below.
        with st.expander("🪄 Auto-discover URLs (ScrapingBee)", expanded=False):
            if _has_scrapingbee_key():
                discover_url = st.text_input(
                    "Crexi search URL",
                    placeholder="https://www.crexi.com/properties?state=GA&types=Retail",
                    help="Paste any Crexi search URL. ScrapingBee scrapes it through a "
                         "residential proxy, we extract listing URLs, you click Add to bulk.",
                    key="sb_discover_url",
                )
                if st.button("🪄 Discover listing URLs", use_container_width=True, key="sb_discover_btn"):
                    if not discover_url.strip().startswith("http"):
                        st.error("Paste a full Crexi URL starting with http(s)://")
                    else:
                        with st.spinner("Fetching via ScrapingBee (15–30s)…"):
                            try:
                                html = fetch_via_scrapingbee(discover_url.strip())
                                found_urls = extract_crexi_listing_urls(html)
                                st.session_state["sb_discovered_urls"] = found_urls
                                if found_urls:
                                    st.success(f"Found {len(found_urls)} listing URLs.")
                                else:
                                    st.warning("ScrapingBee returned HTML but no Crexi listing "
                                               "URLs matched. The selector may need tuning, or "
                                               "Crexi blocked even ScrapingBee.")
                            except Exception as exc:
                                st.error(f"ScrapingBee failed: {exc}")
                _disc = st.session_state.get("sb_discovered_urls") or []
                if _disc:
                    st.caption(f"**Discovered ({len(_disc)}):**")
                    st.code("\n".join(_disc[:10]) + ("\n…" if len(_disc) > 10 else ""), language="text")
                    if st.button("➕ Add discovered URLs to Bulk URLs field", use_container_width=True, key="sb_add_disc_btn"):
                        existing = st.session_state.get("sb_bulk_urls", "")
                        merged = "\n".join(filter(None, [existing.strip()] + _disc))
                        st.session_state["sb_bulk_urls"] = merged
                        st.session_state["sb_discovered_urls"] = []
                        st.rerun()
            else:
                st.info(
                    "**SCRAPINGBEE_API_KEY not set.** Sign up for the free tier "
                    "(1,000 credits/month, no credit card) at [scrapingbee.com]"
                    "(https://www.scrapingbee.com), then add the key to "
                    "**Streamlit Cloud Settings → Secrets** as:\n\n"
                    "```toml\nSCRAPINGBEE_API_KEY = \"YOUR_KEY\"\n```"
                )

        bulk_urls_in = st.text_area(
            "Bulk Crexi URLs (one per line)",
            placeholder="https://www.crexi.com/properties/2287401/...",
            help="The working live path. Paste manually, or use the auto-discover above. "
                 "Each URL ≈ $0.04 via Apify.",
            key="sb_bulk_urls",
            height=100,
        )
        max_props = st.slider(
            "Max properties (search-mode fallback)", 5, 200, DEFAULT_MAX_PROPERTIES, 5,
            key="sb_max_props",
        )

        _parsed_urls = [
            u.strip() for u in bulk_urls_in.splitlines()
            if u.strip().startswith("http")
        ]
        if _parsed_urls:
            st.caption(f"**{len(_parsed_urls)} URLs** queued · est. **\\${len(_parsed_urls) * 0.04:,.2f}**")
        else:
            st.caption(f"Search mode · max {max_props} · est. **\\${max_props * 0.04:,.2f}**")

        run_btn = st.button(
            "🔄 Fetch live deals from Crexi",
            type="primary",
            use_container_width=True,
            disabled=not _token_ok(),
            help=None if _token_ok() else "Set APIFY_TOKEN in .env or st.secrets first",
            key="sb_run_btn",
        )

        with st.expander("…or load an existing Apify dataset", expanded=False):
            ds_id_in = st.text_input("Dataset id", value=APIFY_DATASET_ID, key="sb_ds_id")
            load_ds_btn = st.button(
                "Load dataset", use_container_width=True,
                disabled=not _token_ok(), key="sb_load_ds_btn",
            )

        st.divider()
        st.caption(f"APIFY_TOKEN: {'✅ set' if _token_ok() else '⚠️ missing'}")
        st.caption(f"SCRAPINGBEE_API_KEY: {'✅ set' if _has_scrapingbee_key() else '⚠️ not set (optional)'}")
        if _USE_APIFY_CLIENT:
            st.caption("Apify backend: ✅ `apify-client`")
        elif _APIFY_CLIENT_AVAILABLE:
            st.caption("Apify backend: 🟢 raw REST (set APIFY_USE_CLIENT=1 to opt in)")
        else:
            st.caption("Apify backend: 🟢 raw REST")

    # Underwriting / rule values are now controlled by the expander at the top
    # of the Deal Screener tab. The sidebar ships defaults; the screener tab
    # merges its widget values in before downstream functions consume `sb`.
    return {
        "actor_id": _actor_meta["id"],
        "actor_label": actor_choice,
        "transaction_type": transaction_type,
        "state_code": state_code,
        "asset_class": None if asset_class == "(any)" else asset_class,
        "crexi_category": None if crexi_cat == "(any)" else crexi_cat,
        "crexi_sub_type": None if crexi_sub == "(any)" else crexi_sub,
        "city_or_county": city_or_county_in.strip(),
        "lease_type": None if lease_type_in == "(any)" else lease_type_in,
        "price_min_m": float(price_min_in),
        "price_max_m": float(price_max_in),
        "cap_min": float(cap_min_in),
        "extra_keywords": search_keywords_in.strip(),
        "search_query_preview": _preview_query,
        "bulk_urls": _parsed_urls,
        "max_props": max_props,
        "run_btn": run_btn,
        "ds_id_in": ds_id_in,
        "load_ds_btn": load_ds_btn,
        # CCIM defaults — overridden by the Screener tab's expander widgets.
        "min_cap": DEFAULT_MIN_CAP,
        "max_price_rule": DEFAULT_MAX_PRICE,
        "hold_years": DEFAULT_HOLD_YEARS,
        "noi_growth": DEFAULT_NOI_GROWTH,
        "exit_cap_delta": DEFAULT_EXIT_CAP_DELTA_BPS,
        "ltv": DEFAULT_LTV,
        "loan_rate": DEFAULT_LOAN_RATE,
        "amort_years": DEFAULT_AMORT_YEARS,
        "discount_rate": DEFAULT_DISCOUNT_RATE,
    }


# ---------- Per-deal rendering ----------

def _investment_for_deal(deal_dict: dict, sb: dict, noi_override: float | None = None) -> dict | None:
    price = _coerce_float(deal_dict.get("asking_price"))
    cap = _coerce_float(deal_dict.get("cap_rate_pct"))
    if not price or not cap:
        return None
    exit_cap = cap + sb["exit_cap_delta"] / 100.0
    return project_investment(
        price=price,
        cap_rate_pct=cap,
        hold_years=sb["hold_years"],
        noi_growth_pct=sb["noi_growth"],
        exit_cap_pct=exit_cap,
        ltv_pct=sb["ltv"],
        loan_rate_pct=sb["loan_rate"],
        amort_years=sb["amort_years"],
        discount_rate_pct=sb["discount_rate"],
        noi_override=noi_override,
    )


def _render_investment_block(inv: dict) -> None:
    cols = st.columns(5)
    cols[0].metric("Equity", f"${inv['equity']:,.0f}")
    cols[1].metric("Loan", f"${inv['loan']:,.0f}")
    cols[2].metric("Year-1 CoC", f"{inv['coc_y1']*100:.2f}%")
    cols[3].metric(
        "Levered IRR",
        f"{inv['levered_irr']*100:.2f}%" if inv["levered_irr"] is not None else "—",
    )
    cols[4].metric("Equity Multiple", f"{inv['equity_multiple']:.2f}x")

    cols2 = st.columns(3)
    cols2[0].metric("Annual Debt Service", f"${inv['annual_ds']:,.0f}")
    cols2[1].metric("Reversion @ exit", f"${inv['reversion']:,.0f}")
    cols2[2].metric(
        "NPV @ discount rate", f"${inv['levered_npv']:,.0f}",
        delta="positive = clears hurdle" if inv["levered_npv"] >= 0 else "below hurdle",
    )

    cf_df = investment_cf_dataframe(inv)
    st.dataframe(
        cf_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "NOI": st.column_config.NumberColumn("NOI", format="$%,.0f"),
            "Debt Service": st.column_config.NumberColumn("Debt Service", format="$%,.0f"),
            "CF After Debt": st.column_config.NumberColumn("CF After Debt", format="$%,.0f"),
        },
    )


def render_command_center(deal: pd.Series, idx: int, sb: dict) -> None:
    price_txt = f"${deal['asking_price']:,.0f}" if pd.notna(deal["asking_price"]) else "—"
    cap_txt = f"{deal['cap_rate_pct']:.2f}% cap" if pd.notna(deal["cap_rate_pct"]) else "no cap"
    verdict_txt = deal.get("verdict") or VERDICT_REVIEW
    header = f"{verdict_txt}  ·  {deal['address']}  ·  {price_txt}  ·  {cap_txt}"

    with st.expander(header, expanded=False):
        # Asset-class-specific advisories (Phase 1 env. + 15-yr depreciation).
        asset_label = (sb.get("asset_class") or "").strip()
        crexi_sub_label = (sb.get("crexi_sub_type") or "").strip()
        if asset_label in PHASE_1_ENV_ASSET_CLASSES or crexi_sub_label in CREXI_SUBTYPE_PHASE_1:
            st.warning(
                "⚠️ **Phase 1 Environmental Audit Required** for Underground Storage Tanks / "
                "Chemical Runoff. Budget $3K–$8K and 2–4 weeks before closing diligence.",
                icon="⚠️",
            )
        if bool(deal.get("15_Yr_Accelerated_Depreciation")):
            st.info(
                "⚡ **IRS Class 57.1 — 15-year accelerated depreciation.** Drives a much higher "
                "after-tax IRR than the 39-year default; coordinate cost-segregation with your CPA.",
                icon="⚡",
            )
        deal_flags = list(deal.get("flags") or [])
        if deal_flags:
            high_margin = [f for f in deal_flags if f in HIGH_MARGIN_TRIGGERS]
            other = [f for f in deal_flags if f not in HIGH_MARGIN_TRIGGERS]
            line = "🚩 **Ancillary signals:** "
            if high_margin:
                line += "**" + ", ".join(high_margin) + "** (high-margin)"
            if other:
                line += ("  ·  " if high_margin else "") + ", ".join(other)
            st.markdown(line)

        left, right = st.columns([3, 2])
        with left:
            st.text_input("Property address", value=deal["address"], key=f"addr_{idx}", disabled=True)
        with right:
            st.text_input(
                "Enter LLC from PropTracer",
                key=f"llc_{idx}",
                placeholder="e.g. CHASE BANK CUMMING HOLDINGS LLC",
            )

        st.markdown("**Document audit handoff**")
        om_url = (deal.get("om_url") or "").strip()
        listing_url = (deal.get("listing_url") or "").strip()
        c_om, c_listing = st.columns(2)
        with c_om:
            if om_url:
                st.link_button(
                    "📄 Download OM for Claude Audit",
                    om_url, type="primary", use_container_width=True,
                )
            else:
                st.caption("No OM URL in Apify payload — open the Crexi listing to grab the OM, "
                           "then upload it in the Analyzer tab.")
        with c_listing:
            if listing_url:
                st.link_button(
                    "🔗 Open on Crexi",
                    listing_url, use_container_width=True,
                )

        st.markdown("**Investment analysis (uses sidebar assumptions)**")
        inv = _investment_for_deal(deal.to_dict(), sb)
        if inv is None:
            st.info("Need both asking price and cap rate to run the projection.")
        else:
            _render_investment_block(inv)

        st.markdown("**Underwriting**")
        xlsx_bytes = build_ccim_workbook(deal.to_dict(), investment=inv)
        st.download_button(
            "🧮 Generate CCIM Excel Model",
            data=xlsx_bytes,
            file_name=f"CCIM_Underwriting_{_slugify(str(deal['address']))}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary",
            use_container_width=True,
            key=f"xlsx_{idx}",
        )


# ---------- Screener tab ----------

def render_screener_tab(sb: dict) -> None:
    # CCIM 101 controls — screening rule + underwriting assumptions.
    # Hidden behind a collapsed expander so the tab isn't cluttered for users
    # who only want the listing table; opens when they need to tune the math.
    with st.expander("⚙️ CCIM rule + underwriting assumptions", expanded=False):
        st.caption(
            "Drives the Action-Required verdict and the per-deal investment "
            "analysis (IRR, NPV, equity multiple). Defaults are sensible starting "
            "points; tune to your acquisition strategy."
        )
        st.markdown("**Action-Required rule**")
        rc1, rc2 = st.columns(2)
        with rc1:
            scr_min_cap = st.number_input(
                "Min cap rate (%)", 0.0, 25.0, DEFAULT_MIN_CAP, 0.1, key="scr_min_cap",
            )
        with rc2:
            scr_max_price_rule = st.number_input(
                "Max asking price ($)", 0, 100_000_000, DEFAULT_MAX_PRICE, 100_000,
                key="scr_max_price_rule",
            )

        st.markdown("**Underwriting assumptions (CCIM-style)**")
        uc1, uc2, uc3, uc4 = st.columns(4)
        with uc1:
            scr_hold = st.number_input("Hold years", 1, 30, DEFAULT_HOLD_YEARS, 1, key="scr_hold")
        with uc2:
            scr_growth = st.number_input(
                "NOI growth (%)", 0.0, 15.0, DEFAULT_NOI_GROWTH, 0.25, key="scr_growth",
            )
        with uc3:
            scr_exit_delta = st.number_input(
                "Exit cap delta (bps)", -200, 500, DEFAULT_EXIT_CAP_DELTA_BPS, 25,
                key="scr_exit_delta",
                help="Exit cap = entry cap + this many bps. 0 = same cap.",
            )
        with uc4:
            scr_ltv = st.number_input("LTV (%)", 0.0, 100.0, DEFAULT_LTV, 1.0, key="scr_ltv")

        uc5, uc6, uc7 = st.columns(3)
        with uc5:
            scr_loan_rate = st.number_input(
                "Loan rate (%)", 0.0, 20.0, DEFAULT_LOAN_RATE, 0.05, key="scr_loan_rate",
            )
        with uc6:
            scr_amort = st.number_input(
                "Amortization (years)", 5, 40, DEFAULT_AMORT_YEARS, 1, key="scr_amort",
            )
        with uc7:
            scr_discount = st.number_input(
                "Discount rate / target yield (%)", 0.0, 30.0, DEFAULT_DISCOUNT_RATE, 0.5,
                key="scr_discount",
            )

    # Overlay the tab's widget values onto the sidebar dict so downstream code
    # (analyze_and_score, _investment_for_deal, etc.) reads the user's choices.
    sb = {
        **sb,
        "min_cap": float(scr_min_cap),
        "max_price_rule": float(scr_max_price_rule),
        "hold_years": int(scr_hold),
        "noi_growth": float(scr_growth),
        "exit_cap_delta": int(scr_exit_delta),
        "ltv": float(scr_ltv),
        "loan_rate": float(scr_loan_rate),
        "amort_years": int(scr_amort),
        "discount_rate": float(scr_discount),
    }

    if sb["run_btn"]:
        bulk_urls = sb.get("bulk_urls") or []
        # Crexi's search parser stumbles on punctuation (`/`, double spaces); strip.
        raw_query = sb["search_query_preview"]
        query = re.sub(r"\s+", " ", re.sub(r"[/]+", " ", raw_query)).strip()

        # Build the exact payload we'll send — exposed via st.info on screen
        # so the user can verify filters before / after the run. Sources:
        # (1) the strategy preset's ASSET_CLASS_CATALOG mapping,
        # (2) the Crexi-native category selectbox.
        _types: list[str] = []
        if sb["asset_class"]:
            _types += ASSET_CLASS_CATALOG.get(sb["asset_class"], [])
        if sb["crexi_category"]:
            _types.append(sb["crexi_category"])
        # Dedupe while preserving order.
        seen: set[str] = set()
        property_types_payload = [t for t in _types if not (t in seen or seen.add(t))]
        city_county = (sb.get("city_or_county") or "").strip()
        # locations: the actor's native field. Crawlerbros only accepts state
        # codes, so we send the state code first; the city/county string is
        # included only as a forward-compat marker (silently ignored today).
        locations_payload: list[str] = [sb["state_code"]]
        if city_county:
            locations_payload.append(city_county)

        actor_id = sb["actor_id"]
        mode_label = f"{len(bulk_urls)} URLs (propertyUrls)" if bulk_urls else f"search query '{query}'"
        with st.status(f"Crexi fetch: actor={actor_id} · {mode_label}", expanded=True) as status:
            try:
                status.write("Starting actor run on Apify…")

                def _on_progress(elapsed: int, run_status: str, items: int) -> None:
                    status.update(label=f"{run_status} · {elapsed}s elapsed · {items} items collected")

                if bulk_urls:
                    rows, run_meta = run_actor_async(
                        APIFY_TOKEN, actor_id,
                        property_urls=bulk_urls,
                        max_items=len(bulk_urls),
                        progress_cb=_on_progress,
                    )
                    _tag = " [PARTIAL]" if run_meta.get("_partial") else ""
                    source_label = (
                        f"bulk-URL fetch{_tag} via {actor_id} "
                        f"({len(bulk_urls)} URLs requested, {len(rows)} returned, run {run_meta.get('id','?')})"
                    )
                else:
                    status.write("⚠️ Search-based fetch is degraded right now (Crexi UI change). "
                                 "Expect 0 results — paste URLs in the sidebar for the working path.")
                    rows, run_meta = run_actor_async(
                        APIFY_TOKEN, actor_id,
                        search_keywords=[query],
                        property_types=property_types_payload or None,
                        locations=locations_payload,
                        max_items=sb["max_props"],
                        progress_cb=_on_progress,
                    )
                    _tag = " [PARTIAL]" if run_meta.get("_partial") else ""
                    source_label = (
                        f"search{_tag} via {actor_id} query='{query}' "
                        f"({len(rows)} items, run {run_meta.get('id','?')})"
                    )

                st.session_state["deals_rows"] = rows
                st.session_state["data_source"] = source_label
                st.session_state["last_payload"] = _build_actor_payload(
                    search_keywords=None if bulk_urls else [query],
                    start_urls=None,
                    property_urls=bulk_urls or None,
                    property_types=property_types_payload or None,
                    locations=None if bulk_urls else locations_payload,
                    max_items=(len(bulk_urls) if bulk_urls else sb["max_props"]),
                    max_search_pages=5,
                )
                save_last_fetch(rows, source_label, query if not bulk_urls else f"BULK ({len(bulk_urls)} URLs)")
                is_partial = bool(run_meta.get("_partial"))
                if rows and is_partial:
                    partial_count = run_meta.get("_partial_count", len(rows))
                    status.update(
                        label=(
                            f"🟡 Partial fetch — Apify run ended early "
                            f"({run_meta.get('status','?')}) after collecting {partial_count} listings. "
                            f"Showing what we got."
                        ),
                        state="complete",
                    )
                elif rows:
                    status.update(label=f"✅ Fetched {len(rows)} GA listings from Crexi.", state="complete")
                else:
                    status.update(
                        label=(
                            f"⚠️ Fetch completed but returned 0 items. "
                            f"{'No matching listings.' if bulk_urls else 'Search actor degraded — switch to Bulk URL mode.'}"
                        ),
                        state="error",
                    )
            except Exception as exc:
                exc_text = str(exc)
                status.update(label=f"❌ {exc_text}", state="error")
                if "TIMED-OUT" in exc_text or "TIMEOUT" in exc_text.upper():
                    st.error(
                        "**Apify run timed out before scraping any results.** The residential "
                        "proxy is slow to handshake with Crexi. Try one of:\n\n"
                        "1. **Drop Max properties to 5** (sidebar slider) — smaller jobs land faster.\n"
                        "2. **Add or change the City/County** in the sidebar — narrower searches "
                        "return fewer pages to scrape.\n"
                        "3. **Wait 60 seconds and click Fetch again** — Apify's residential proxy "
                        "pool sometimes recovers between attempts.\n"
                        "4. **Try a different asset class hint** — some Crexi queries (rare niches) "
                        "return a slow-loading detail page."
                    )
                else:
                    st.error(
                        "Try lowering 'Max properties' to 5, or narrow the query "
                        "(add a city/county) to reduce scrape time."
                    )

    if sb["load_ds_btn"] and sb["ds_id_in"].strip():
        with st.spinner(f"Loading dataset {sb['ds_id_in']}…"):
            try:
                rows = fetch_dataset_items(APIFY_TOKEN, sb["ds_id_in"].strip())
                st.session_state["deals_rows"] = rows
                st.session_state["data_source"] = f"dataset {sb['ds_id_in'].strip()} ({len(rows)} items)"
                st.success(f"Loaded {len(rows)} items from dataset.")
            except Exception as exc:
                st.error(f"Dataset load failed: {exc}")

    rows = st.session_state.get("deals_rows")
    disk_loaded = False
    age_min = None

    # If session was reset (Cloud rebuild, tab close), recover the last
    # successful live fetch from disk. NO sample-data fallback per spec.
    if not rows:
        last = load_last_fetch()
        if last and last.get("rows"):
            rows = last["rows"]
            st.session_state["deals_rows"] = rows
            st.session_state["data_source"] = last.get("source", "disk cache")
            disk_loaded = True
            try:
                age_min = (datetime.utcnow() - datetime.fromisoformat(last["ts"].rstrip("Z"))).total_seconds() / 60
            except Exception:
                age_min = None

    # Hard empty state — no mock data, no fallback. Prompt the user to fetch.
    if not rows:
        with st.container(border=True):
            st.subheader("📡 No live deals loaded")
            st.warning(
                "**Crexi search-based scraping is degraded right now.** Two different Apify "
                "actors (skootle, cypherai) both return 0 items for Crexi search results. "
                "The direct-URL fetch path still works. **Workflow:**\n\n"
                "1. Open **crexi.com** in another tab and run your search there (their human search works).\n"
                "2. Copy the URLs of the listings you want to analyze.\n"
                "3. Paste them into **Bulk URL mode** in the sidebar (one per line).\n"
                "4. Click 🔄 Fetch — each URL ≈ $0.04, returns full data (cap rate, SF, broker, etc.)."
            )
            st.markdown(
                "This dashboard shows **only live Apify data** — there is no demo / mock dataset. "
                "Configure the search in the sidebar and click 🔄 to pull listings."
            )
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("**Connected actor**")
                st.code(DEFAULT_ACTOR_ID, language="text")
                st.caption(
                    "skootle returns 50+ fields per listing including "
                    "`address`, `capRatePct`, `squareFootageNum`, `description`, "
                    "and broker contact. The Action-Required / GO verdict logic "
                    "depends on these fields."
                )
            with c2:
                st.markdown("**Current search**")
                st.code(sb["search_query_preview"], language="text")
                st.caption(
                    f"State: locked to **{DEFAULT_STATE_CODE}** · "
                    f"City/County: **{sb.get('city_or_county') or '(any)'}** · "
                    f"Asset: **{sb.get('asset_class') or '(any)'}** · "
                    f"Max: **{sb['max_props']}**"
                )
            st.markdown(
                "👉 Click **🔄 Fetch live deals from Crexi** in the sidebar to start a run "
                f"(estimated cost ~${sb['max_props'] * 0.04:.2f} at Free tier)."
            )
        return

    df = normalize_rows(rows)

    # Hard sale/lease filter — Apify's `transactionType` is reliable on the rows
    # that have it ("sale" or "lease"). Empty values are kept (some skootle
    # responses don't populate it).
    txn = sb.get("transaction_type")
    if txn == "For Sale":
        df = df[df["transaction_type_native"].isin(["sale", ""])]
    elif txn == "For Lease":
        df = df[df["transaction_type_native"].isin(["lease", ""])]

    df = analyze_and_score(df, sb)

    if disk_loaded:
        ago = f"{age_min:.0f} min ago" if age_min is not None else "earlier"
        c_left, c_right = st.columns([5, 1])
        c_left.info(f"📂 Restored last live fetch from disk cache ({ago}). "
                    f"Click 🔄 in the sidebar to pull fresh data.")
        if c_right.button("Clear cache", help="Discard cached data and return to empty state"):
            clear_last_fetch()
            st.session_state["deals_rows"] = None
            st.rerun()
    else:
        st.success(f"🟢 Live data: {st.session_state['data_source']}", icon="✅")

    # Per Product Owner spec: show the exact Apify run_input payload that was
    # last sent. Lets you verify filters (state, locations, propertyTypes, etc.)
    # without round-tripping to the Apify console.
    last_payload = st.session_state.get("last_payload")
    if last_payload:
        with st.expander("🛰️ Last Apify run_input payload (verify filters)", expanded=False):
            st.json(last_payload)

    # ----- Top metrics -----
    go_count = int((df["verdict"] == VERDICT_GO).sum()) if not df.empty else 0
    action_count = int((df["verdict"] == VERDICT_ACTION).sum()) if not df.empty else 0
    tax_alpha_count = int(df["15_Yr_Accelerated_Depreciation"].sum()) if "15_Yr_Accelerated_Depreciation" in df.columns else 0
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Total deals", len(df))
    m2.metric("🟢 GO (Tax-Alpha)", go_count)
    m3.metric("🟢 Action Required", action_count)
    m4.metric("⚡ 15-yr eligible", tax_alpha_count)
    m5.metric("Avg cap rate", f"{df['cap_rate_pct'].mean():.2f}%" if not df.empty and df["cap_rate_pct"].notna().any() else "—")

    # ----- Live filters (apply to the table below) -----
    with st.container(border=True):
        st.markdown("**Filters** — apply to the table and chart below in real time")
        types_available = sorted([t for t in df["property_type"].dropna().unique().tolist() if t])
        f1, f2, f3 = st.columns(3)
        with f1:
            type_filter = st.multiselect(
                "Asset type", types_available, default=[], key="scr_type_filter",
                placeholder="(all types)",
            )
            verdict_filter = st.multiselect(
                "Verdict",
                [VERDICT_GO, VERDICT_ACTION, VERDICT_REVIEW],
                default=[], key="scr_status_filter",
                placeholder="(all verdicts)",
            )
        with f2:
            prices = df["asking_price"].dropna()
            if len(prices) >= 1:
                p_min, p_max = int(prices.min()), max(int(prices.max()), int(prices.min()) + 1)
                price_range = st.slider(
                    "Asking price ($)", p_min, p_max, (p_min, p_max),
                    step=max((p_max - p_min) // 50, 50_000),
                    key="scr_price_range",
                )
            else:
                price_range = None
                st.caption("No price data to filter.")
        with f3:
            caps = df["cap_rate_pct"].dropna()
            if len(caps) >= 1:
                c_min, c_max = float(caps.min()), max(float(caps.max()), float(caps.min()) + 0.1)
                cap_range = st.slider(
                    "Cap rate (%)", c_min, c_max, (c_min, c_max), step=0.1, format="%.2f",
                    key="scr_cap_range",
                )
            else:
                cap_range = None
                st.caption("No cap data to filter.")

    fdf = df.copy()
    if type_filter:
        fdf = fdf[fdf["property_type"].isin(type_filter)]
    if verdict_filter:
        fdf = fdf[fdf["verdict"].isin(verdict_filter)]
    if price_range:
        fdf = fdf[fdf["asking_price"].fillna(-1).between(price_range[0], price_range[1])]
    if cap_range:
        fdf = fdf[fdf["cap_rate_pct"].fillna(-1).between(cap_range[0], cap_range[1])]
    fdf = fdf.reset_index(drop=True)

    st.subheader(f"All deals · {len(fdf)} after filter")
    # Reorder columns so the verdict / tax-alpha / flags signals show first.
    display_cols = [c for c in [
        "address", "property_type", "sub_class",
        "asking_price", "cap_rate_pct", "square_footage",
        "verdict", "15_Yr_Accelerated_Depreciation", "flags",
        "om_url", "listing_url",
    ] if c in fdf.columns]
    display_df = fdf[display_cols] if display_cols else fdf
    event = st.dataframe(
        display_df,
        use_container_width=True, hide_index=True,
        on_select="rerun", selection_mode="single-row",
        key="scr_table",
        column_config={
            "asking_price": st.column_config.NumberColumn("Asking Price", format="$%,.0f"),
            "cap_rate_pct": st.column_config.NumberColumn("Cap Rate %", format="%.2f"),
            "square_footage": st.column_config.NumberColumn("SF", format="%,d"),
            "om_url": st.column_config.LinkColumn("OM"),
            "listing_url": st.column_config.LinkColumn("Listing"),
            "verdict": st.column_config.TextColumn("Verdict"),
            "property_type": st.column_config.TextColumn("Type"),
            "sub_class": st.column_config.TextColumn(
                "Sub-Class",
                help="Parsed from listing title + description + native subtype tags. "
                     "Falls back to the main asset class when no keywords match.",
            ),
            "15_Yr_Accelerated_Depreciation": st.column_config.CheckboxColumn(
                "⚡ 15-yr", help="IRS Class 57.1 — qualifies for 15-year accelerated depreciation",
                disabled=True,
            ),
            "flags": st.column_config.ListColumn("🚩 Ancillary flags", width="medium"),
        },
    )

    # ----- Inline deep-dive for the row the user clicked -----
    selected_rows = getattr(event.selection, "rows", []) if hasattr(event, "selection") else []
    if selected_rows and not fdf.empty:
        sel = fdf.iloc[selected_rows[0]]
        with st.container(border=True):
            verdict_txt = sel.get("verdict") or VERDICT_REVIEW
            st.subheader(f"🔍 Selected: {sel['address']}  ·  {verdict_txt}")

            asset_label = (sb.get("asset_class") or "").strip()
            if asset_label in PHASE_1_ENV_ASSET_CLASSES:
                st.warning(
                    "⚠️ **Phase 1 Environmental Audit Required** for Underground Storage Tanks / "
                    "Chemical Runoff. Budget $3K–$8K and 2–4 weeks before closing diligence.",
                    icon="⚠️",
                )
            if bool(sel.get("15_Yr_Accelerated_Depreciation")):
                st.info(
                    "⚡ **IRS Class 57.1 — 15-year accelerated depreciation.** Drives a much higher "
                    "after-tax IRR than the 39-year default; coordinate cost-segregation with your CPA.",
                    icon="⚡",
                )
            sel_flags = list(sel.get("flags") or [])
            if sel_flags:
                hi = [f for f in sel_flags if f in HIGH_MARGIN_TRIGGERS]
                lo = [f for f in sel_flags if f not in HIGH_MARGIN_TRIGGERS]
                msg = "🚩 **Ancillary signals:** "
                if hi:
                    msg += "**" + ", ".join(hi) + "** (high-margin)"
                if lo:
                    msg += ("  ·  " if hi else "") + ", ".join(lo)
                st.markdown(msg)

            head = st.columns(4)
            head[0].metric("Asking Price", f"${sel['asking_price']:,.0f}" if pd.notna(sel["asking_price"]) else "—")
            head[1].metric("Cap Rate", f"{sel['cap_rate_pct']:.2f}%" if pd.notna(sel["cap_rate_pct"]) else "—")
            head[2].metric("SF", f"{int(sel['square_footage']):,}" if pd.notna(sel["square_footage"]) else "—")
            head[3].metric(
                "Price/SF",
                f"${sel['asking_price']/sel['square_footage']:,.0f}"
                if pd.notna(sel["asking_price"]) and pd.notna(sel["square_footage"]) and sel["square_footage"] else "—",
            )
            link_cols = st.columns(2)
            if sel.get("listing_url"):
                link_cols[0].link_button("🔗 Open on Crexi", sel["listing_url"], use_container_width=True)
            if sel.get("om_url"):
                link_cols[1].link_button("📄 OM", sel["om_url"], type="primary", use_container_width=True)

            inv = _investment_for_deal(sel.to_dict(), sb)
            if inv:
                _render_investment_block(inv)
                xlsx_bytes = build_ccim_workbook(sel.to_dict(), investment=inv)
                st.download_button(
                    "🧮 Generate CCIM Excel Model",
                    data=xlsx_bytes,
                    file_name=f"CCIM_Underwriting_{_slugify(str(sel['address']))}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="scr_sel_xlsx",
                )
            else:
                st.info("Need both asking price and cap rate to run the investment projection.")
    else:
        st.caption("👆 Click any row in the table to see a deep dive with investment analysis + CCIM Excel.")

    # ----- Scatter chart -----
    chart_df = fdf.dropna(subset=["asking_price", "cap_rate_pct"]).copy()
    if len(chart_df) >= 2:
        with st.expander("📈 Cap rate vs asking price (interactive chart)", expanded=False):
            chart_df["sf_for_size"] = chart_df["square_footage"].fillna(chart_df["square_footage"].dropna().median() if chart_df["square_footage"].notna().any() else 10000)
            chart = (
                alt.Chart(chart_df)
                .mark_circle(opacity=0.75, stroke="white", strokeWidth=1)
                .encode(
                    x=alt.X("asking_price:Q", title="Asking Price ($)", axis=alt.Axis(format="$,.0f")),
                    y=alt.Y("cap_rate_pct:Q", title="Cap Rate (%)", scale=alt.Scale(zero=False)),
                    color=alt.Color("verdict:N", legend=alt.Legend(title="Verdict")),
                    size=alt.Size("sf_for_size:Q", title="SF", scale=alt.Scale(range=[80, 600]), legend=None),
                    tooltip=[
                        alt.Tooltip("address:N", title="Address"),
                        alt.Tooltip("property_type:N", title="Type"),
                        alt.Tooltip("asking_price:Q", title="Price", format="$,.0f"),
                        alt.Tooltip("cap_rate_pct:Q", title="Cap %", format=".2f"),
                        alt.Tooltip("square_footage:Q", title="SF", format=",.0f"),
                        alt.Tooltip("verdict:N", title="Verdict"),
                    ],
                )
                .properties(height=320)
                .interactive()
            )
            st.altair_chart(chart, use_container_width=True)

    # ----- Action Required Command Center (kept as an alternative drill-in) -----
    action_df = fdf[fdf["verdict"].isin([VERDICT_GO, VERDICT_ACTION])].reset_index(drop=True)
    with st.expander(
        f"🟢 Due Diligence Command Center — GO + Action Required ({len(action_df)} deals)",
        expanded=False,
    ):
        if action_df.empty:
            st.info("No deals carry a GO / Action Required verdict under the current filters.")
        else:
            for i, deal in action_df.iterrows():
                render_command_center(deal, i, sb)


# ---------- Analyzer tab ----------

def render_analyzer_tab(sb: dict) -> None:
    st.write(
        "Paste a Crexi link **or** upload an OM PDF **or** type the deal in by hand. "
        "Anything we can parse pre-fills the manual fields below — you confirm and click Analyze."
    )

    if PdfReader is None:
        st.warning("`pypdf` not installed — PDF upload is disabled. `pip install pypdf` to enable.")

    col_url, col_pdf = st.columns([3, 2])
    with col_url:
        url_in = st.text_input(
            "Crexi listing URL",
            key="analyzer_url",
            placeholder="https://www.crexi.com/properties/2287401/georgia-chase-bank-cumming-ga",
            help="Click 'Fetch from Crexi' to auto-populate the form via Apify (~$0.04 per deal).",
        )
        fetch_crexi_btn = st.button(
            "🔎 Fetch from Crexi (auto-fill via Apify)",
            disabled=not (url_in and _token_ok()),
            use_container_width=True,
            help="Runs the skootle actor with propertyUrls=[this URL], maxItems=1.",
        )
    with col_pdf:
        pdf_file = st.file_uploader(
            "OM / Flyer PDF (optional)",
            type=["pdf"],
            key="analyzer_pdf",
            disabled=PdfReader is None,
        )

    if fetch_crexi_btn and url_in:
        with st.spinner("Fetching this listing from Crexi via Apify…"):
            try:
                rows = run_actor_sync(
                    APIFY_TOKEN, DEFAULT_ACTOR_ID,
                    property_urls=[url_in], max_items=1,
                )
                if rows:
                    st.session_state["analyzer_fetched"] = rows[0]
                    st.success(f"Fetched listing. {len(rows[0])} fields populated.")
                else:
                    st.warning("Apify returned 0 items for this URL.")
            except Exception as exc:
                st.error(f"Apify fetch failed: {exc}")

    fetched = st.session_state.get("analyzer_fetched") or {}
    parsed: dict[str, Any] = {}
    if fetched:
        parsed = {
            "price": _coerce_float(_first_present(fetched, ("askingPriceUsd", "askingPrice", "price"))),
            "cap_rate": _coerce_float(_first_present(fetched, ("capRatePct", "capRate", "cap_rate"))),
            "sf": _coerce_int(_first_present(fetched, ("squareFootageNum", "squareFootage", "buildingSqft", "square_footage"))),
            "noi": _coerce_float(_first_present(fetched, ("noiUsd", "netOperatingIncome"))),
        }
        parsed = {k: v for k, v in parsed.items() if v}
        title_or_addr = fetched.get("title") or fetched.get("address") or ""
        if parsed or title_or_addr:
            populated = ", ".join(f"{k}={v}" for k, v in parsed.items()) or "(metadata only)"
            st.caption(f"From Apify fetch: {populated}  ·  {title_or_addr}")
        if fetched.get("askingPriceUsd") is None and fetched.get("askingPrice") is None:
            st.info("⚠️ Crexi has gated the price for this listing — contact the broker "
                    f"({fetched.get('brokerName','—')}, {fetched.get('brokerPhone','—')}). "
                    "Enter price manually below to run the underwriting.")

    pdf_text = ""
    if pdf_file is not None and PdfReader is not None:
        pdf_text = extract_pdf_text(pdf_file.getvalue())
        pdf_parsed = parse_om_fields(pdf_text)
        # PDF parse can override Apify if values are present; user can still edit.
        for k, v in pdf_parsed.items():
            parsed.setdefault(k, v) if k in parsed else parsed.update({k: v})
        if pdf_parsed:
            st.success(f"Parsed {len(pdf_parsed)} field(s) from PDF: {', '.join(pdf_parsed.keys())}")
        else:
            st.info("PDF uploaded but no fields auto-detected — fill them in below.")

    slug_hint = extract_crexi_slug_hint(url_in)
    if fetched:
        composed = _compose_address(fetched)
        default_addr = composed if composed != "—" else (fetched.get("title") or slug_hint or "")
    else:
        default_addr = slug_hint or ""

    st.markdown("#### Deal details")
    c1, c2 = st.columns(2)
    with c1:
        address = st.text_input("Address / nickname", value=default_addr, key="an_address",
                                placeholder="123 Main St, Atlanta, GA 30309")
        price = st.number_input(
            "Asking price ($)", min_value=0.0, value=float(parsed.get("price") or 0.0),
            step=50_000.0, key="an_price",
        )
        cap = st.number_input(
            "Cap rate (%)", min_value=0.0, max_value=25.0, value=float(parsed.get("cap_rate") or 0.0),
            step=0.05, key="an_cap",
        )
    with c2:
        sf = st.number_input(
            "Square footage", min_value=0, value=int(parsed.get("sf") or 0), step=500, key="an_sf",
        )
        noi_override = st.number_input(
            "NOI override ($, optional)",
            min_value=0.0, value=float(parsed.get("noi") or 0.0), step=10_000.0,
            help="Use the OM's stated NOI; leave 0 to derive NOI as price × cap.",
            key="an_noi",
        )
        om_url_in = st.text_input("OM/Flyer URL (optional)", value="", key="an_om",
                                  placeholder="https://…/offering-memorandum.pdf")

    st.markdown("#### Override sidebar assumptions (optional)")
    c3, c4, c5 = st.columns(3)
    with c3:
        hold = st.number_input("Hold years", 1, 30, sb["hold_years"], 1, key="an_hold")
    with c4:
        growth = st.number_input("NOI growth (%)", 0.0, 15.0, sb["noi_growth"], 0.25, key="an_growth")
    with c5:
        exit_cap_in = st.number_input(
            "Exit cap (%)", 0.0, 20.0, max(0.0, (cap or 7.0) + sb["exit_cap_delta"] / 100.0), 0.05,
            key="an_exit",
        )
    c6, c7, c8, c9 = st.columns(4)
    with c6:
        ltv_in = st.number_input("LTV (%)", 0.0, 100.0, sb["ltv"], 1.0, key="an_ltv")
    with c7:
        loan_rate_in = st.number_input("Loan rate (%)", 0.0, 20.0, sb["loan_rate"], 0.05, key="an_loanr")
    with c8:
        amort_in = st.number_input("Amort (years)", 5, 40, sb["amort_years"], 1, key="an_amort")
    with c9:
        disc_in = st.number_input("Discount rate (%)", 0.0, 30.0, sb["discount_rate"], 0.5, key="an_disc")

    analyze = st.button("🚀 Analyze deal", type="primary", use_container_width=True)
    if not analyze:
        if pdf_text:
            with st.expander("PDF text preview (first 2000 chars)"):
                st.text(pdf_text[:2000])
        return

    if not price or not cap:
        st.error("Need both an asking price and a cap rate (either typed in or parsed from the PDF).")
        return

    inv = project_investment(
        price=price, cap_rate_pct=cap, hold_years=int(hold), noi_growth_pct=float(growth),
        exit_cap_pct=float(exit_cap_in), ltv_pct=float(ltv_in), loan_rate_pct=float(loan_rate_in),
        amort_years=int(amort_in), discount_rate_pct=float(disc_in),
        noi_override=noi_override or None,
    )

    st.divider()
    st.subheader(f"🔬 {address or 'Untitled deal'}")
    head_cols = st.columns(4)
    head_cols[0].metric("Asking Price", f"${price:,.0f}")
    head_cols[1].metric("Cap Rate", f"{cap:.2f}%")
    head_cols[2].metric("Square Footage", f"{int(sf):,}" if sf else "—")
    head_cols[3].metric("Price / SF", f"${price/sf:,.0f}" if sf else "—")

    if url_in:
        prop_id = extract_crexi_property_id(url_in)
        st.caption(f"Source: {url_in}" + (f"  ·  property_id={prop_id}" if prop_id else ""))

    if om_url_in:
        st.link_button("📄 Download OM for Claude Audit", om_url_in,
                       type="primary", use_container_width=False)

    st.markdown("#### Investment analysis")
    _render_investment_block(inv)

    deal_dict = {
        "address": address or slug_hint or "Untitled deal",
        "asking_price": price,
        "cap_rate_pct": cap,
        "square_footage": int(sf) if sf else None,
        "listing_url": url_in,
        "om_url": om_url_in,
    }
    xlsx_bytes = build_ccim_workbook(deal_dict, investment=inv)
    st.download_button(
        "🧮 Generate CCIM Excel Model (with Investment Analysis sheet)",
        data=xlsx_bytes,
        file_name=f"CCIM_Underwriting_{_slugify(deal_dict['address'])}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="secondary",
        use_container_width=True,
    )

    if pdf_text:
        with st.expander("PDF text preview (first 2000 chars)"):
            st.text(pdf_text[:2000])


# ---------- AI Analyst (Claude API) ----------

CRE_ANALYST_SYSTEM_PROMPT = """You are a senior commercial real estate (CRE) \
investment analyst with 20 years of experience underwriting acquisitions for \
institutional investors. You think like an institutional buyer:

- Lease structure first, asset second. Tenant credit quality matters more than day-1 yield.
- Replacement cost is the floor; cap rate is the ceiling. Always ask: would we build this for less?
- For NNN: who owns the roof, structure, and parking lot? Read the lease before the broker's marketing.
- For multifamily: in-place rent vs market rent. What's the rent growth runway? What are real OpEx ratios for this asset class in this market?
- Tax-advantaged sub-classes (gas stations, car washes — IRS Class 57.1, 15-yr accelerated depreciation) need Phase 1 environmental on USTs / chemical runoff.
- High-margin ancillary revenue (COAM gaming, lottery, unbranded fuel margin, vending) can carry a deal — but verify it's transferable and re-licensable.

When responding:
1. Lead with the verdict and the single most important reason — no preamble.
2. Be specific. Cite numbers from the data provided. Don't speculate.
3. If critical data is missing, say what's missing and how to get it. List it as the first question to ask the broker.
4. Cite specific OM passages when relevant (e.g. "per OM: 'lease expires 9/30/2040'").
5. Use markdown: headers, bullet points, **bold** for key terms and numbers.

Never:
- Hedge with "it depends" without stating the dependency explicitly.
- Use generic CRE platitudes ("location is important").
- Make up data not in the OM or context."""

CRE_ACTION_PROMPTS: dict[str, str] = {
    "summarize": (
        "Write a tight 200-word executive summary of this deal. Cover: what it is, "
        "the headline financials, the tenant/lease story, and your one-line take. "
        "No bullets — flowing prose."
    ),
    "risks": (
        "Identify the **top 5 specific risks** in this deal. For each: name the risk in bold, "
        "explain it in 1–2 sentences with citation to the data or OM, and rate severity "
        "(High / Medium / Low). Skip generic risks that apply to all CRE."
    ),
    "questions": (
        "Generate the **top 7 questions** I should ask the listing broker before signing an LOI. "
        "Order by importance. Each question should be specific to this deal, citing the data or "
        "OM passage that prompts it. Skip questions whose answers are already in the materials."
    ),
    "verdict": (
        "Give your **Go / No-Go / Conditional Go** verdict on this deal. Lead with the verdict "
        "in bold. Then give the 3 strongest reasons in bullet form. Then 1–2 sentences on what "
        "would change your call. Keep total response under 250 words."
    ),
    "full": (
        "Produce a 1-page institutional underwriting memo with these exact sections:\n"
        "## Verdict (Go / No-Go / Conditional Go, with 1-sentence reason)\n"
        "## Deal Snapshot (3–5 bullets: address, asset, key financials)\n"
        "## Lease & Tenant Analysis (4–6 bullets)\n"
        "## Returns Profile (entry yield, downside, comps)\n"
        "## Top 5 Risks (ranked)\n"
        "## Top 5 Broker Questions (ranked)\n"
        "## Bottom Line (2–3 sentences)"
    ),
}


def _deal_context_block(ctx: dict[str, Any]) -> str:
    def _f(key, fmt=str, default="—"):
        v = ctx.get(key)
        if v is None or v == "" or (isinstance(v, float) and pd.isna(v)):
            return default
        try:
            return fmt(v)
        except Exception:
            return str(v)
    return (
        "# Deal Context\n"
        f"- Address: {_f('address')}\n"
        f"- Asking Price: ${_f('asking_price', lambda x: f'{float(x):,.0f}')}\n"
        f"- Cap Rate: {_f('cap_rate_pct', lambda x: f'{float(x):.2f}%')}\n"
        f"- Square Footage: {_f('square_footage', lambda x: f'{int(x):,}')}\n"
        f"- Asset Class: {_f('property_type')}\n"
        f"- Sub-Class: {_f('sub_class')}\n"
        f"- Listing URL: {_f('listing_url')}\n"
        f"- Verdict (heuristic): {_f('verdict')}\n"
        f"- IRS 15-yr eligible: {_f('15_Yr_Accelerated_Depreciation')}\n"
        f"- Ancillary flags: {_f('flags', lambda x: ', '.join(x) if x else '—')}\n"
    )


def stream_claude_analysis(action: str, deal_ctx: dict[str, Any], om_text: str):
    """Stream Claude's response for the chosen action.

    Uses prompt caching on the system prompt AND the OM text — so successive
    actions on the same deal hit cache and cost ~1/10 the first request.
    Yields text chunks suitable for st.write_stream.
    """
    if not _ANTHROPIC_AVAILABLE or Anthropic is None:
        raise RuntimeError("anthropic package not installed")
    if not ANTHROPIC_API_KEY or ANTHROPIC_API_KEY == ANTHROPIC_PLACEHOLDER:
        raise RuntimeError("ANTHROPIC_API_KEY not configured (.env or st.secrets)")

    client = Anthropic(api_key=ANTHROPIC_API_KEY)
    action_prompt = CRE_ACTION_PROMPTS.get(action, "Analyze this deal.")
    context_block = _deal_context_block(deal_ctx)

    user_content: list[dict[str, Any]] = [
        {
            "type": "text",
            "text": context_block,
        },
    ]
    if om_text and om_text.strip():
        # Truncate aggressively-long OMs; Sonnet has plenty of context but we
        # want to keep the first action snappy.
        clipped = om_text[:80_000]
        user_content.append({
            "type": "text",
            "text": f"\n\n# Offering Memorandum (extracted text)\n\n{clipped}",
            "cache_control": {"type": "ephemeral"},
        })
    user_content.append({
        "type": "text",
        "text": f"\n\n# Task\n\n{action_prompt}",
    })

    with client.messages.stream(
        model=ANTHROPIC_MODEL,
        max_tokens=4096,
        system=[
            {
                "type": "text",
                "text": CRE_ANALYST_SYSTEM_PROMPT,
                "cache_control": {"type": "ephemeral"},
            },
        ],
        messages=[{"role": "user", "content": user_content}],
    ) as stream:
        for text in stream.text_stream:
            yield text


def _has_anthropic_key() -> bool:
    return bool(ANTHROPIC_API_KEY) and ANTHROPIC_API_KEY != ANTHROPIC_PLACEHOLDER


def render_ai_analyst_tab(sb: dict) -> None:
    st.subheader("🤖 AI Analyst — Claude underwrites the deal")
    st.caption(
        f"Powered by **{ANTHROPIC_MODEL}** with prompt caching. "
        f"~$0.05–0.15 for the first action on a new OM; ~$0.01–0.03 for each "
        f"follow-up action on the same OM (cache hit)."
    )

    if not _ANTHROPIC_AVAILABLE:
        st.error(
            "`anthropic` SDK not installed in this environment. Locally: "
            "`pip install anthropic`. On Streamlit Cloud, it'll install on the "
            "next rebuild after requirements.txt is updated."
        )
        return
    if not _has_anthropic_key():
        st.warning(
            "**ANTHROPIC_API_KEY is not set.**\n\n"
            "- **Locally:** add `ANTHROPIC_API_KEY=sk-ant-...` to `.env`\n"
            "- **Streamlit Cloud:** Settings → Secrets → add "
            "`ANTHROPIC_API_KEY = \"sk-ant-...\"`\n\n"
            "Get a key at [console.anthropic.com](https://console.anthropic.com/account/keys)."
        )
        return

    # ---------- Source picker ----------
    src_tabs = st.tabs([
        "📋 From Screener",
        "📄 Upload OM PDF",
        "✏️ Paste OM Text",
    ])
    deal_ctx: dict[str, Any] = {}
    om_text = ""

    with src_tabs[0]:
        rows = st.session_state.get("deals_rows") or []
        if not rows:
            st.info("No deals loaded in the Screener yet. Fetch some first.")
        else:
            df = normalize_rows(rows)
            df = analyze_and_score(df, sb)
            if df.empty:
                st.info("Loaded deals didn't pass the GA filter.")
            else:
                options: dict[str, int] = {
                    f"{i+1}. {row['address']} — ${row['asking_price']:,.0f}"
                    if pd.notna(row.get("asking_price")) else
                    f"{i+1}. {row['address']}"
                    : i
                    for i, row in df.iterrows()
                }
                choice = st.selectbox("Pick a deal", list(options.keys()), key="analyst_pick")
                idx = options[choice]
                deal_ctx = df.iloc[idx].to_dict()
                cols = st.columns(4)
                cols[0].metric("Asking Price", f"${deal_ctx.get('asking_price', 0):,.0f}" if pd.notna(deal_ctx.get("asking_price")) else "—")
                cols[1].metric("Cap Rate", f"{deal_ctx.get('cap_rate_pct', 0):.2f}%" if pd.notna(deal_ctx.get("cap_rate_pct")) else "—")
                cols[2].metric("SF", f"{int(deal_ctx.get('square_footage', 0)):,}" if pd.notna(deal_ctx.get("square_footage")) else "—")
                cols[3].metric("Verdict", deal_ctx.get("verdict", "—"))

    with src_tabs[1]:
        if PdfReader is None:
            st.warning("`pypdf` not installed — PDF upload unavailable.")
        else:
            pdf_file = st.file_uploader(
                "Upload an Offering Memorandum (PDF)",
                type=["pdf"], key="analyst_pdf",
            )
            if pdf_file is not None:
                om_text = extract_pdf_text(pdf_file.getvalue())
                if om_text:
                    st.success(f"Extracted {len(om_text):,} characters from PDF.")
                    with st.expander("PDF text preview (first 2000 chars)"):
                        st.text(om_text[:2000])
                    parsed = parse_om_fields(om_text)
                    if parsed:
                        st.caption(f"Auto-parsed: {', '.join(f'{k}={v}' for k, v in parsed.items())}")
                        deal_ctx.setdefault("asking_price", parsed.get("price"))
                        deal_ctx.setdefault("cap_rate_pct", parsed.get("cap_rate"))
                        deal_ctx.setdefault("square_footage", parsed.get("sf"))
                else:
                    st.warning("Couldn't extract text — the PDF may be image-only (scanned).")

    with src_tabs[2]:
        manual_addr = st.text_input("Property address / nickname", key="analyst_addr_manual")
        c1, c2, c3 = st.columns(3)
        with c1:
            manual_price = st.number_input("Asking Price ($)", min_value=0.0, value=0.0, step=50_000.0, key="analyst_price_manual")
        with c2:
            manual_cap = st.number_input("Cap Rate (%)", min_value=0.0, max_value=25.0, value=0.0, step=0.05, key="analyst_cap_manual")
        with c3:
            manual_sf = st.number_input("Square Footage", min_value=0, value=0, step=500, key="analyst_sf_manual")
        pasted = st.text_area("Paste OM text here", height=240, key="analyst_paste")
        if pasted.strip():
            om_text = pasted
            deal_ctx = {
                "address": manual_addr,
                "asking_price": manual_price or None,
                "cap_rate_pct": manual_cap or None,
                "square_footage": manual_sf or None,
                "property_type": "",
            }

    # ---------- Action buttons ----------
    if not deal_ctx and not om_text:
        st.info("👆 Pick a source above to enable analysis.")
        return

    st.divider()
    st.markdown("**Choose an analysis. Output streams below in real time.**")
    bc = st.columns(5)
    triggered_action: str | None = None
    if bc[0].button("📝 Summarize", use_container_width=True, key="ai_sum"):
        triggered_action = "summarize"
    if bc[1].button("⚠️ Top 5 Risks", use_container_width=True, key="ai_risk"):
        triggered_action = "risks"
    if bc[2].button("❓ Broker Q's", use_container_width=True, key="ai_qs"):
        triggered_action = "questions"
    if bc[3].button("✅ Go / No-Go", use_container_width=True, key="ai_v"):
        triggered_action = "verdict"
    if bc[4].button("📊 Full Memo", use_container_width=True, type="primary", key="ai_full"):
        triggered_action = "full"

    if triggered_action:
        st.divider()
        st.markdown(f"### Claude — {triggered_action.replace('_',' ').title()}")
        try:
            output = st.write_stream(stream_claude_analysis(triggered_action, deal_ctx, om_text))
        except Exception as exc:
            st.error(f"Analysis failed: {exc}")
            return
        # Offer the result as a download / copy.
        st.download_button(
            "💾 Download analysis (.md)",
            data=output if isinstance(output, str) else "".join(output),
            file_name=f"claude_{triggered_action}_{_slugify(str(deal_ctx.get('address','deal')))}.md",
            mime="text/markdown",
            use_container_width=False,
        )


# ---------- Main ----------

def main() -> None:
    st.set_page_config(page_title="CRE Deal Screener", layout="wide", page_icon="🏢")
    st.title("CRE Deal Screener — Georgia")
    st.caption("Screener + Single-Deal Analyzer. Levered DCF math via numpy_financial (same as tvm.py).")
    st.session_state.setdefault("deals_rows", None)
    st.session_state.setdefault("data_source", "sample")

    sb = render_sidebar()
    tab_screener, tab_analyzer = st.tabs(["📊 Screener", "🔬 Single-Deal Analyzer"])
    with tab_screener:
        render_screener_tab(sb)
    with tab_analyzer:
        render_analyzer_tab(sb)


if __name__ == "__main__":
    main()
